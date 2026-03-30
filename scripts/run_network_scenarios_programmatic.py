#!/usr/bin/env python3
"""Run network topology comparison using the programmatic Network API.

Usage:
    python scripts/run_network_scenarios_programmatic.py

This script builds scenarios entirely in Python — no YAML files needed.
Useful for parameter sweeps (pipe diameter, length, temperature levels).
"""

from __future__ import annotations

import sys
import time
from pathlib import Path
from typing import Any, Callable, Dict, List

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from calion.network import Network


# ---------------------------------------------------------------------------
# Shared base configuration
# ---------------------------------------------------------------------------

def _base_network(name: str, solver: str = "appsi_highs") -> Network:
    """Create a base Network with common components (no topology yet)."""
    net = Network(dt_h=1.0, solver=solver, name=name)
    net.add_heat_pump("HP1", capacity_mw=80.0, cop=3.5)
    net.add_generator("Boiler", cap_th_mw=150.0, fuel_bus="gas", efficiency=0.90)
    net.set_fuel("gas", price_eur_mwh=45.0, ef_kg_per_mwh_fuel=200.0)
    net.set_grid(max_import_mw=200.0)
    net.set_costs(co2_price_eur_per_t=100.0)
    return net


def _load_timeseries(net: Network, hours: int = 168) -> Network:
    """Load or synthesize time series for the network.

    Tries to load from Import_Data.xlsx; falls back to synthetic data.
    """
    xlsx = PROJECT_ROOT / "Import_Data.xlsx"
    if xlsx.exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(xlsx), data_only=True)
            ws = wb.active
            headers = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]

            # Find columns
            demand_col = next((h for h in headers if "rmebedarf" in h or "demand" in h.lower()), None)
            price_col = next((h for h in headers if "Price" in h or "preis" in h.lower()), None)
            co2_col = next((h for h in headers if "CO2" in h), None)

            # Read rows
            demand, prices, co2 = [], [], []
            for row in ws.iter_rows(min_row=2, max_row=2 + hours - 1):
                cells = {h: c.value for h, c in zip(headers, row)}
                if demand_col:
                    demand.append(float(cells.get(demand_col, 0) or 0))
                if price_col:
                    prices.append(float(cells.get(price_col, 50) or 50))
                if co2_col:
                    co2.append(float(cells.get(co2_col, 400) or 400))

            if demand:
                net.set_demand(demand[:hours])
            if prices:
                net.set_electricity_price(prices[:hours])
            if co2:
                net.set_grid_co2(co2[:hours])
            return net
        except Exception:
            pass

    # Fallback: synthetic winter week
    import math
    demand = [
        30 + 20 * math.sin(2 * math.pi * h / 24 - math.pi / 2)  # 10–50 MW daily cycle
        for h in range(hours)
    ]
    prices = [
        50 + 30 * math.sin(2 * math.pi * h / 24)  # 20–80 EUR/MWh
        for h in range(hours)
    ]
    co2 = [400.0] * hours

    net.set_demand(demand)
    net.set_electricity_price(prices)
    net.set_grid_co2(co2)
    return net


# ---------------------------------------------------------------------------
# Topology builders
# ---------------------------------------------------------------------------

def build_copperplate(solver: str = "appsi_highs") -> Network:
    """Single-node (no network) — copperplate model."""
    net = _base_network("copperplate", solver)
    return _load_timeseries(net)


def build_star(solver: str = "appsi_highs") -> Network:
    """Star: plant → 3 consumers (radial, no junctions)."""
    net = _base_network("star", solver)
    net.set_network_temps(supply_temp_c=90, return_temp_c=55, ground_temp_c=10)
    net.set_network_physics(heat_loss=True)

    net.add_node("plant", "producer", assets=["HP1", "Boiler"])
    net.add_node("north", "consumer", demand_fraction=0.40)
    net.add_node("south", "consumer", demand_fraction=0.35)
    net.add_node("east", "consumer", demand_fraction=0.25)

    net.add_pipe("plant_north", "plant", "north", length_m=2000, diameter_mm=600)
    net.add_pipe("plant_south", "plant", "south", length_m=1500, diameter_mm=500)
    net.add_pipe("plant_east", "plant", "east", length_m=3000, diameter_mm=500)

    return _load_timeseries(net)


def build_tree(solver: str = "appsi_highs") -> Network:
    """Tree: plant → junction → branches (realistic urban layout)."""
    net = _base_network("tree", solver)
    net.set_network_temps(supply_temp_c=90, return_temp_c=55, ground_temp_c=10)
    net.set_network_physics(heat_loss=True)

    net.add_node("plant", "producer", assets=["HP1", "Boiler"])
    net.add_node("jct_main", "junction")
    net.add_node("jct_west", "junction")
    net.add_node("north", "consumer", demand_fraction=0.30)
    net.add_node("south", "consumer", demand_fraction=0.25)
    net.add_node("west_a", "consumer", demand_fraction=0.25)
    net.add_node("west_b", "consumer", demand_fraction=0.20)

    net.add_pipe("trunk", "plant", "jct_main", length_m=1000, diameter_mm=900)
    net.add_pipe("to_north", "jct_main", "north", length_m=2000, diameter_mm=500)
    net.add_pipe("to_south", "jct_main", "south", length_m=1500, diameter_mm=500)
    net.add_pipe("to_west_jct", "jct_main", "jct_west", length_m=800, diameter_mm=600)
    net.add_pipe("west_a", "jct_west", "west_a", length_m=600, diameter_mm=500)
    net.add_pipe("west_b", "jct_west", "west_b", length_m=900, diameter_mm=400)

    return _load_timeseries(net)


def build_ring(solver: str = "appsi_highs") -> Network:
    """Ring: closed loop — redundant supply paths."""
    net = _base_network("ring", solver)
    net.set_network_temps(supply_temp_c=90, return_temp_c=55, ground_temp_c=10)
    net.set_network_physics(heat_loss=True)

    net.add_node("plant", "producer", assets=["HP1", "Boiler"])
    net.add_node("north", "consumer", demand_fraction=0.40)
    net.add_node("east", "consumer", demand_fraction=0.25)
    net.add_node("south", "consumer", demand_fraction=0.35)

    net.add_pipe("plant_north", "plant", "north", length_m=1500, diameter_mm=900)
    net.add_pipe("north_east", "north", "east", length_m=1200, diameter_mm=700)
    net.add_pipe("east_south", "east", "south", length_m=1000, diameter_mm=600)
    net.add_pipe("south_plant", "south", "plant", length_m=1800, diameter_mm=300)

    return _load_timeseries(net)


def build_long_pipe(solver: str = "appsi_highs") -> Network:
    """Linear chain: plant → C1 → C2 → C3 (worst-case end-of-line losses)."""
    net = _base_network("long_pipe", solver)
    net.set_network_temps(supply_temp_c=95, return_temp_c=55, ground_temp_c=10)
    net.set_network_physics(heat_loss=True)

    net.add_node("plant", "producer", assets=["HP1", "Boiler"])
    net.add_node("c1", "consumer", demand_fraction=0.50)
    net.add_node("c2", "consumer", demand_fraction=0.30)
    net.add_node("c3", "consumer", demand_fraction=0.20)

    net.add_pipe("p1", "plant", "c1", length_m=3000, diameter_mm=800)
    net.add_pipe("p2", "c1", "c2", length_m=3000, diameter_mm=600)
    net.add_pipe("p3", "c2", "c3", length_m=3000, diameter_mm=400)

    return _load_timeseries(net)


# ---------------------------------------------------------------------------
# Parameter sweep helpers
# ---------------------------------------------------------------------------

def diameter_sweep(
    base_builder: Callable[..., Network],
    pipe_id: str,
    diameters_mm: List[float],
    solver: str = "appsi_highs",
) -> List[Dict[str, Any]]:
    """Run a pipe diameter sensitivity analysis.

    Parameters
    ----------
    base_builder : callable
        Function that returns a configured Network (e.g., ``build_star``).
    pipe_id : str
        Which pipe to vary.
    diameters_mm : list of float
        Diameter values to sweep.

    Returns
    -------
    list of dict
        One result dict per diameter value.
    """
    results = []
    for d in diameters_mm:
        net = base_builder(solver=solver)
        if pipe_id in net._pipes:
            net._pipes[pipe_id]["diameter_mm"] = d
        net.name = f"d={d}mm"

        cfg = net._build_config()
        info = {
            "diameter_mm": d,
            "total_pipe_length_m": sum(p.get("length_m", 0) for p in net._pipes.values()),
            "num_pipes": len(net._pipes),
        }

        # Run optimization
        try:
            result = net.optimize()
            info.update({
                "status": "optimal" if result.is_optimal else "other",
                "total_cost_eur": result.total_cost(),
            })
        except Exception as e:
            info.update({"status": "error", "error": str(e)})

        results.append(info)
    return results


def temperature_sweep(
    base_builder: Callable[..., Network],
    supply_temps: List[float],
    solver: str = "appsi_highs",
) -> List[Dict[str, Any]]:
    """Run a supply temperature sensitivity analysis."""
    results = []
    for t_supply in supply_temps:
        net = base_builder(solver=solver)
        net.set_network_temps(supply_temp_c=t_supply, return_temp_c=55, ground_temp_c=10)
        net.name = f"T_s={t_supply}C"

        info = {"supply_temp_c": t_supply}
        try:
            result = net.optimize()
            info.update({
                "status": "optimal" if result.is_optimal else "other",
                "total_cost_eur": result.total_cost(),
            })
        except Exception as e:
            info.update({"status": "error", "error": str(e)})

        results.append(info)
    return results


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

BUILDERS = {
    "copperplate": build_copperplate,
    "star": build_star,
    "tree": build_tree,
    "ring": build_ring,
    "long_pipe": build_long_pipe,
}


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Programmatic network topology comparison")
    parser.add_argument("--scenarios", nargs="*", default=None)
    parser.add_argument("--solver", default="appsi_highs")
    parser.add_argument("--hours", type=int, default=168, help="Number of hours to simulate")
    parser.add_argument("--sweep", choices=["diameter", "temperature"], default=None)
    args = parser.parse_args()

    scenarios = args.scenarios or list(BUILDERS.keys())

    if args.sweep == "diameter":
        print("Running diameter sweep on star topology...")
        results = diameter_sweep(build_star, "plant_north", [200, 250, 300, 350, 400, 500], args.solver)
        for r in results:
            print(f"  DN{r['diameter_mm']}: {r.get('status', '?')} — {r.get('total_cost_eur', 'N/A')}")
        return

    if args.sweep == "temperature":
        print("Running supply temperature sweep on star topology...")
        results = temperature_sweep(build_star, [70, 80, 90, 100, 110, 120], args.solver)
        for r in results:
            print(f"  T_s={r['supply_temp_c']}°C: {r.get('status', '?')} — {r.get('total_cost_eur', 'N/A')}")
        return

    # Standard topology comparison
    print(f"\nBuilding and running {len(scenarios)} scenarios...\n")

    results = []
    for name in scenarios:
        builder = BUILDERS.get(name)
        if not builder:
            print(f"  Unknown scenario: {name}")
            continue

        net = builder(solver=args.solver)
        print(f"  [{name}] {len(net._nodes)} nodes, {len(net._pipes)} pipes")

        t0 = time.time()
        try:
            result = net.optimize()
            elapsed = time.time() - t0
            results.append({
                "scenario": name,
                "status": "optimal" if result.is_optimal else "other",
                "total_cost_eur": result.total_cost(),
                "co2_t": result.co2_total_t(),
                "grid_cost_eur": result.grid_cost(),
                "fuel_cost_eur": result.fuel_cost(),
                "co2_cost_eur": result.co2_cost(),
                "nodes": len(net._nodes),
                "pipes": len(net._pipes),
                "pipe_length_m": sum(p.get("length_m", 0) for p in net._pipes.values()),
                "solve_time_s": round(elapsed, 1),
            })
            print(f"    -> {results[-1]['status']} | {results[-1]['total_cost_eur']:,.0f} EUR | {elapsed:.1f}s")
        except Exception as e:
            elapsed = time.time() - t0
            print(f"    -> ERROR: {e} ({elapsed:.1f}s)")
            results.append({
                "scenario": name,
                "status": "error",
                "error": str(e),
                "nodes": len(net._nodes),
                "pipes": len(net._pipes),
            })

    # Summary
    if results:
        print("\n" + "=" * 100)
        print(f"{'Scenario':<15} {'Status':<9} {'Total [EUR]':>13} {'Grid [EUR]':>12} {'Fuel [EUR]':>12} {'CO2 [EUR]':>11} {'CO2 [t]':>9} {'Pipes':>6} {'Length':>8}")
        print("-" * 100)
        for r in results:
            print(
                f"{r['scenario']:<15} {r['status']:<9} "
                f"{r.get('total_cost_eur', 0):>13,.0f} "
                f"{r.get('grid_cost_eur', 0):>12,.0f} "
                f"{r.get('fuel_cost_eur', 0):>12,.0f} "
                f"{r.get('co2_cost_eur', 0):>11,.0f} "
                f"{r.get('co2_t', 0):>9,.0f} "
                f"{r.get('pipes', 0):>6} "
                f"{r.get('pipe_length_m', 0):>8,.0f}"
            )
        print("=" * 100)


if __name__ == "__main__":
    main()
