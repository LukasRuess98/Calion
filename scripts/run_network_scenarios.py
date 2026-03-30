#!/usr/bin/env python3
"""Run and compare multiple network topology scenarios.

Usage:
    python scripts/run_network_scenarios.py
    python scripts/run_network_scenarios.py --scenarios star ring tree
    python scripts/run_network_scenarios.py --output results/network_comparison.csv

Each scenario is a YAML config in configs/scenarios/network_*.yaml.
Results are compared side-by-side: total cost, heat losses, HP share, etc.
"""

from __future__ import annotations

import argparse
import csv
import logging
import sys
import time
from pathlib import Path
from typing import Any, Dict, List, Optional

# Project root
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

SCENARIO_DIR = PROJECT_ROOT / "configs" / "scenarios"

# Available network topology scenarios
TOPOLOGY_SCENARIOS = {
    "copperplate": "test_3node_network.yaml",   # no pipes (single-node)
    "star":        "network_star.yaml",           # radial: plant → 3 consumers
    "ring":        "network_ring.yaml",           # loop with redundant paths
    "tree":        "network_tree.yaml",           # branched hierarchy with junctions
    "long_pipe":   "network_long_pipe.yaml",      # serial chain (worst-case losses)
    "5node":       None,                          # level2_5node (unified format)
}


def _load_config(scenario_name: str) -> Dict[str, Any]:
    """Load and merge a scenario config file."""
    from energis.config.merge import load_and_merge

    fname = TOPOLOGY_SCENARIOS.get(scenario_name)
    if fname is None:
        # Check examples dir
        fpath = PROJECT_ROOT / "configs" / "examples" / f"level2_5node.yaml"
        if not fpath.exists():
            raise FileNotFoundError(f"No config for scenario '{scenario_name}'")
    else:
        fpath = SCENARIO_DIR / fname
        if not fpath.exists():
            raise FileNotFoundError(f"Config not found: {fpath}")

    config = load_and_merge(str(fpath))
    return config


def _count_pipes(config: Dict[str, Any]) -> int:
    """Count pipes in a config (various formats)."""
    # Unified format
    net = config.get("network", {})
    pipes = net.get("pipes", {})
    if isinstance(pipes, dict):
        return len(pipes)
    if isinstance(pipes, list):
        return len(pipes)
    # Legacy format
    tn = config.get("thermal_network", {})
    if tn:
        networks = tn.get("networks", {})
        total = 0
        for n in (networks.values() if isinstance(networks, dict) else networks):
            p = n.get("pipes", {})
            total += len(p) if isinstance(p, (dict, list)) else 0
        return total
    return 0


def _count_nodes(config: Dict[str, Any]) -> int:
    """Count nodes in a config."""
    net = config.get("network", {})
    nodes = net.get("nodes", {})
    if isinstance(nodes, dict):
        return len(nodes)
    if isinstance(nodes, list):
        return len(nodes)
    return 0


def _total_pipe_length(config: Dict[str, Any]) -> float:
    """Sum up all pipe lengths in metres."""
    net = config.get("network", {})
    pipes = net.get("pipes", {})
    total = 0.0
    if isinstance(pipes, dict):
        for p in pipes.values():
            total += float(p.get("length_m", 0))
    elif isinstance(pipes, list):
        for p in pipes:
            total += float(p.get("length_m", 0))
    return total


def run_scenario(scenario_name: str, config: Dict[str, Any]) -> Dict[str, Any]:
    """Run a single scenario and return summary metrics."""
    from energis.run.solver import _solve_scenario
    from energis.utils.timeseries import TimeSeriesTable

    log.info(f"--- Running scenario: {scenario_name} ---")

    # Extract run settings
    run_cfg = config.get("run", {})
    dt_h = run_cfg.get("dt_h", 1.0)
    solver = run_cfg.get("solver", "appsi_highs")

    # Load time series
    site = config.get("site", {})
    xlsx_path = PROJECT_ROOT / site.get("input_xlsx", "Import_Data.xlsx")
    if not xlsx_path.exists():
        log.error(f"Data file not found: {xlsx_path}")
        return {"scenario": scenario_name, "status": "MISSING_DATA"}

    # Load data via the standard pipeline
    try:
        from energis.run.utilities.timeseries_utils import _load_and_prepare_table
        table = _load_and_prepare_table(config, str(PROJECT_ROOT))
    except Exception as e:
        log.warning(f"Could not load table via pipeline: {e}")
        log.info("Attempting direct Excel load...")
        try:
            import openpyxl
            from datetime import datetime
            wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
            ws = wb.active
            headers = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
            data = {h: [] for h in headers}
            for row in ws.iter_rows(min_row=2):
                for h, cell in zip(headers, row):
                    data[h].append(cell.value)
            n = len(data[headers[0]])
            index = [datetime(2024, 1, 1, 0, 0) for _ in range(n)]  # placeholder
            table = TimeSeriesTable(index=index, columns=headers, data=data)
        except Exception as e2:
            log.error(f"Failed to load data: {e2}")
            return {"scenario": scenario_name, "status": "LOAD_ERROR", "error": str(e2)}

    # Solve
    t0 = time.time()
    try:
        result = _solve_scenario(table, config, dt_h, solver)
        elapsed = time.time() - t0
    except Exception as e:
        elapsed = time.time() - t0
        log.error(f"Solver failed after {elapsed:.1f}s: {e}")
        return {
            "scenario": scenario_name,
            "status": "SOLVER_ERROR",
            "error": str(e),
            "solve_time_s": round(elapsed, 2),
        }

    # Extract metrics
    costs = result.costs or {}
    summary = result.summary or {}

    metrics = {
        "scenario": scenario_name,
        "status": str(result.solver.get("termination_condition", "unknown")),
        "solve_time_s": round(elapsed, 2),
        "total_cost_eur": round(costs.get("total_cost_eur", 0), 2),
        "energy_cost_eur": round(costs.get("energy_cost_eur", 0), 2),
        "co2_cost_eur": round(costs.get("co2_cost_eur", 0), 2),
        "fuel_cost_eur": round(costs.get("fuel_cost_eur", 0), 2),
        "total_co2_t": round(summary.get("total_co2_t", 0), 2),
        "hp_share_pct": round(summary.get("hp_share_pct", 0), 1),
        "network_loss_mwh": round(summary.get("network_loss_mwh", 0), 2),
        "num_nodes": _count_nodes(config),
        "num_pipes": _count_pipes(config),
        "total_pipe_length_m": _total_pipe_length(config),
    }

    log.info(
        f"  Status: {metrics['status']} | "
        f"Cost: {metrics['total_cost_eur']:,.0f} EUR | "
        f"Losses: {metrics['network_loss_mwh']:.1f} MWh | "
        f"Time: {metrics['solve_time_s']:.1f}s"
    )
    return metrics


def print_comparison(results: List[Dict[str, Any]]) -> None:
    """Print a side-by-side comparison table."""
    if not results:
        return

    print("\n" + "=" * 90)
    print("NETWORK TOPOLOGY COMPARISON")
    print("=" * 90)

    # Header
    header_fields = [
        ("Scenario", 15),
        ("Status", 10),
        ("Cost [EUR]", 14),
        ("CO2 [t]", 10),
        ("HP [%]", 8),
        ("Loss [MWh]", 11),
        ("Pipes", 6),
        ("Length [m]", 10),
        ("Time [s]", 8),
    ]
    header = " | ".join(f"{name:>{width}}" for name, width in header_fields)
    print(header)
    print("-" * len(header))

    for r in results:
        row = [
            f"{r.get('scenario', '?'):>15}",
            f"{r.get('status', '?'):>10}",
            f"{r.get('total_cost_eur', 0):>14,.0f}",
            f"{r.get('total_co2_t', 0):>10,.1f}",
            f"{r.get('hp_share_pct', 0):>8.1f}",
            f"{r.get('network_loss_mwh', 0):>11,.1f}",
            f"{r.get('num_pipes', 0):>6}",
            f"{r.get('total_pipe_length_m', 0):>10,.0f}",
            f"{r.get('solve_time_s', 0):>8.1f}",
        ]
        print(" | ".join(row))

    print("=" * 90)

    # Highlight key findings
    solved = [r for r in results if "optimal" in str(r.get("status", "")).lower()]
    if len(solved) >= 2:
        cheapest = min(solved, key=lambda r: r.get("total_cost_eur", float("inf")))
        lossiest = max(solved, key=lambda r: r.get("network_loss_mwh", 0))
        greenest = min(solved, key=lambda r: r.get("total_co2_t", float("inf")))
        print(f"\n  Cheapest:     {cheapest['scenario']} ({cheapest['total_cost_eur']:,.0f} EUR)")
        print(f"  Highest loss: {lossiest['scenario']} ({lossiest['network_loss_mwh']:,.1f} MWh)")
        print(f"  Lowest CO2:   {greenest['scenario']} ({greenest['total_co2_t']:,.1f} t)")


def save_csv(results: List[Dict[str, Any]], output_path: Path) -> None:
    """Save results to CSV."""
    if not results:
        return
    output_path.parent.mkdir(parents=True, exist_ok=True)
    keys = results[0].keys()
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        writer.writerows(results)
    log.info(f"Results saved to {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Run network topology comparison")
    parser.add_argument(
        "--scenarios",
        nargs="*",
        default=None,
        help="Scenario names to run (default: all available)",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Output CSV path (optional)",
    )
    parser.add_argument(
        "--solver",
        type=str,
        default=None,
        help="Override solver for all scenarios",
    )
    args = parser.parse_args()

    # Determine which scenarios to run
    scenario_names = args.scenarios or ["copperplate", "star", "ring", "tree", "long_pipe"]

    log.info(f"Running {len(scenario_names)} network scenarios: {scenario_names}")

    results = []
    for name in scenario_names:
        try:
            config = _load_config(name)
            if args.solver:
                config.setdefault("run", {})["solver"] = args.solver
            metrics = run_scenario(name, config)
            results.append(metrics)
        except FileNotFoundError as e:
            log.warning(f"Skipping {name}: {e}")
            results.append({"scenario": name, "status": "NOT_FOUND"})
        except Exception as e:
            log.error(f"Error in {name}: {e}")
            results.append({"scenario": name, "status": "ERROR", "error": str(e)})

    print_comparison(results)

    if args.output:
        save_csv(results, Path(args.output))


if __name__ == "__main__":
    main()
