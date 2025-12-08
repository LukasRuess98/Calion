#!/usr/bin/env python3
"""
Create Excel template for thermal network configuration.

This script generates a structured Excel file with 6 sheets that can be
filled out and converted to YAML using the thermal_network_excel_parser.

Usage:
    python scripts/create_thermal_network_template.py --output data/my_network.xlsx
    python scripts/create_thermal_network_template.py --help
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any, Dict, List

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False
    print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)


def create_network_sheet(wb: Workbook) -> None:
    """Create the Netzwerk (Network) sheet with basic parameters."""
    ws = wb.create_sheet("Netzwerk", 0)

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Headers
    ws['A1'] = 'Parameter'
    ws['B1'] = 'Wert'
    ws['C1'] = 'Einheit'
    ws['D1'] = 'Beschreibung'

    for col in ['A1', 'B1', 'C1', 'D1']:
        ws[col].fill = header_fill
        ws[col].font = header_font

    # Example data
    data = [
        ['Name', 'Musterhausen', '', 'Name des Wärmenetzes'],
        ['T_Vorlauf_nom', 90, '°C', 'Nominale Vorlauftemperatur'],
        ['T_Ruecklauf_nom', 50, '°C', 'Nominale Rücklauftemperatur'],
        ['p_nominal', 6, 'bar', 'Nominaler Netzdruck'],
        ['Netztyp', 'district_heating', '', 'district_heating | building_network'],
        ['Optimierungsmodus', 'cost', '', 'cost | emissions | exergy'],
    ]

    for i, row_data in enumerate(data, start=2):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if j == 1:  # Parameter column
                cell.font = Font(bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 50


def create_producers_sheet(wb: Workbook) -> None:
    """Create the Erzeuger (Producers) sheet."""
    ws = wb.create_sheet("Erzeuger", 1)

    # Header styling
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Headers
    headers = [
        'ID', 'Typ', 'Bestand?', 'Investition?', 'Q_nom_MW', 'Q_options_MW',
        'CAPEX_€_kW', 'OPEX_fix_€_kW_a', 'OPEX_var_€_MWh',
        'Wirkungsgrad', 'COP', 'Brennstoff', 'Vorlauf_Knoten', 'Ruecklauf_Knoten'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Example rows with comments
    examples = [
        ['Kessel_1', 'boiler', 'ja', 'nein', 15.0, '', 300, 15, 25, 0.95, '', 'gas', 'N1', 'N1_R'],
        ['CHP_1', 'chp', 'ja', 'nein', 8.0, '', 1200, 30, 20, 0.85, '', 'gas', 'N1', 'N1_R'],
        ['WP_1', 'heat_pump', 'nein', 'ja', '', '5.0,10.0,15.0', 800, 20, 5, '', 3.5, 'electricity', 'N2', 'N2_R'],
    ]

    # Add comment header
    ws['A2'] = '# Beispiele (löschen und eigene Daten eintragen):'
    ws['A2'].font = Font(italic=True, color="808080")

    for i, row_data in enumerate(examples, start=3):
        for j, value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=value)

    # Column widths
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15


def create_storage_sheet(wb: Workbook) -> None:
    """Create the Speicher (Storage) sheet."""
    ws = wb.create_sheet("Speicher", 2)

    # Header styling
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Headers
    headers = [
        'ID', 'Typ', 'Bestand?', 'Investition?', 'Kapazitaet_MWh', 'Kapazitaet_options_MWh',
        'T_max_C', 'T_min_C', 'Verlustrate_1_h', 'CAPEX_€_kWh', 'OPEX_fix_€_kWh_a',
        'Vorlauf_Knoten', 'Ruecklauf_Knoten'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Example rows
    examples = [
        ['Speicher_1', 'hot_water_tank', 'ja', 'nein', 30.0, '', 95, 40, 0.001, 50, 2, 'N1', 'N1_R'],
        ['PTES_1', 'ptes', 'nein', 'ja', '', '50,100,200', 90, 30, 0.0005, 30, 1, 'N1', 'N1_R'],
    ]

    ws['A2'] = '# Beispiele (löschen und eigene Daten eintragen):'
    ws['A2'].font = Font(italic=True, color="808080")

    for i, row_data in enumerate(examples, start=3):
        for j, value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=value)

    # Column widths
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15


def create_pipes_sheet(wb: Workbook) -> None:
    """Create the Rohre (Pipes) sheet."""
    ws = wb.create_sheet("Rohre", 3)

    # Header styling
    header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Headers
    headers = [
        'ID', 'Von_Knoten', 'Zu_Knoten', 'Laenge_m', 'Bestand?', 'Investition?',
        'DN_fix', 'DN_options', 'CAPEX_€_m', 'Zustand'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Example rows
    examples = [
        ['P1', 'gaskessel_01', 'junction_ctr', 100, 'ja', 'nein', 'DN200', '', 500, 'good'],
        ['P2', 'chp_01', 'junction_ctr', 150, 'ja', 'nein', 'DN150', '', 500, 'good'],
        ['P3', 'junction_ctr', 'altstadt', 5000, 'ja', 'nein', 'DN200', '', 500, 'good'],
        ['P4', 'wp_neu', 'junction_ctr', 500, 'nein', 'ja', '', 'DN150,DN200,DN250', 600, ''],
    ]

    ws['A2'] = '# NUR Vorlaufrohre! Rücklauf wird automatisch generiert.'
    ws['A2'].font = Font(italic=True, color="808080")
    ws['A3'] = '# Beispiele (löschen und eigene Daten eintragen):'
    ws['A3'].font = Font(italic=True, color="808080")

    for i, row_data in enumerate(examples, start=4):
        for j, value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=value)

    # Column widths
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15


def create_consumers_sheet(wb: Workbook) -> None:
    """Create the Verbraucher (Consumers) sheet."""
    ws = wb.create_sheet("Verbraucher", 4)

    # Header styling
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Headers
    headers = [
        'ID', 'Knoten', 'Lastgang', 'Spitzenlast_MW', 'Jahresbedarf_MWh'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Example rows
    examples = [
        ['Altstadt', 'altstadt', 'heat_demand_altstadt', 12.0, 45000],
        ['Neustadt', 'neustadt', 'heat_demand_neustadt', 8.0, 30000],
    ]

    ws['A2'] = '# Lastgang muss in Sheet "Zeitreihen" als Spalte existieren'
    ws['A2'].font = Font(italic=True, color="808080")
    ws['A3'] = '# Beispiele (löschen und eigene Daten eintragen):'
    ws['A3'].font = Font(italic=True, color="808080")

    for i, row_data in enumerate(examples, start=4):
        for j, value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=value)

    # Column widths
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20


def create_timeseries_sheet(wb: Workbook) -> None:
    """Create the Zeitreihen (Timeseries) sheet."""
    ws = wb.create_sheet("Zeitreihen", 5)

    # Header styling
    header_fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Headers
    headers = [
        'Zeitstempel', 'heat_demand_altstadt', 'heat_demand_neustadt',
        'Strompreis_€_MWh', 'Gaspreis_€_MWh', 'Aussentemperatur_C'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Example rows (first few hours)
    ws['A2'] = '2023-01-01 00:00:00'
    ws['B2'] = 8.5
    ws['C2'] = 6.2
    ws['D2'] = 65.2
    ws['E2'] = 35.0
    ws['F2'] = -2.5

    ws['A3'] = '2023-01-01 01:00:00'
    ws['B3'] = 8.2
    ws['C3'] = 6.0
    ws['D3'] = 62.1
    ws['E3'] = 35.0
    ws['F3'] = -3.1

    ws['A4'] = '2023-01-01 02:00:00'
    ws['B4'] = 7.9
    ws['C4'] = 5.8
    ws['D4'] = 58.5
    ws['E4'] = 35.0
    ws['F4'] = -3.8

    ws['A5'] = '# ... (8760 Zeilen für Gesamtjahr)'
    ws['A5'].font = Font(italic=True, color="808080")

    # Column widths
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20


def create_instructions_sheet(wb: Workbook) -> None:
    """Create an instructions sheet."""
    ws = wb.create_sheet("Anleitung", 6)

    # Title
    ws['A1'] = 'Thermal Network Excel Template - Anleitung'
    ws['A1'].font = Font(size=16, bold=True, color="4472C4")

    # Instructions
    instructions = [
        '',
        'VERWENDUNG:',
        '1. Füllen Sie die 6 Sheets aus (Netzwerk, Erzeuger, Speicher, Rohre, Verbraucher, Zeitreihen)',
        '2. Speichern Sie die Excel-Datei',
        '3. Konvertieren Sie zu YAML mit dem thermal_network_excel_parser',
        '',
        'WICHTIGE HINWEISE:',
        '',
        'SHEET "Netzwerk":',
        '- Grundlegende Netzwerk-Parameter',
        '- Vorlauf-/Rücklauftemperaturen, Druck, Optimierungsmodus',
        '',
        'SHEET "Erzeuger":',
        '- Wärmeerzeuger (Kessel, Wärmepumpen, BHKW, etc.)',
        '- WICHTIG: Genau EINS von "Bestand?" oder "Investition?" muss "ja" sein',
        '- Bestand = ja: Komponente existiert bereits (Q_nom ausfüllen)',
        '- Investition = ja: Neue Komponente (Q_options ausfüllen, z.B. "5.0,10.0,15.0")',
        '',
        'SHEET "Speicher":',
        '- Thermische Speicher (optional)',
        '- Gleiche Logik wie Erzeuger: Bestand ODER Investition',
        '',
        'SHEET "Rohre":',
        '- NUR Vorlaufrohre eintragen!',
        '- Rücklaufrohre werden automatisch generiert',
        '- DN_fix für Bestandsrohre (z.B. "DN200")',
        '- DN_options für Investitionsrohre (z.B. "DN150,DN200,DN250")',
        '',
        'SHEET "Verbraucher":',
        '- Wärmeabnehmer mit Lastgang-Referenz',
        '- Lastgang muss als Spalte in "Zeitreihen" existieren',
        '',
        'SHEET "Zeitreihen":',
        '- ALLE Zeitreihen in einem Sheet',
        '- Mindestens 8760 Zeilen für Gesamtjahr',
        '- Spalten: Zeitstempel, Lastgänge, Preise, Temperaturen, etc.',
        '',
        'KONVERTIERUNG:',
        'python -c "from energis.utils.thermal_network_excel_parser import ThermalNetworkExcelParser; \\',
        '           parser = ThermalNetworkExcelParser(\'my_network.xlsx\'); \\',
        '           parser.save_yaml(\'my_network.scenario.yaml\')"',
        '',
        'DOKUMENTATION:',
        '- docs/excel_import_feature.md',
        '- docs/brownfield_quickstart_guide.md',
    ]

    for i, text in enumerate(instructions, start=2):
        ws[f'A{i}'] = text
        if text.startswith('SHEET') or text in ['VERWENDUNG:', 'WICHTIGE HINWEISE:', 'KONVERTIERUNG:', 'DOKUMENTATION:']:
            ws[f'A{i}'].font = Font(bold=True, color="70AD47")

    # Column width
    ws.column_dimensions['A'].width = 100


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='Create Excel template for thermal network configuration',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s --output data/my_network.xlsx
  %(prog)s --output scenarios/stadtbach.xlsx --name Stadtbach
        """
    )

    parser.add_argument(
        '--output', '-o',
        type=str,
        required=True,
        help='Output Excel file path (e.g., data/my_network.xlsx)'
    )

    parser.add_argument(
        '--name', '-n',
        type=str,
        default='Musterhausen',
        help='Network name (default: Musterhausen)'
    )

    args = parser.parse_args()

    # Create output directory if needed
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Create workbook
    print(f"Creating Excel template: {output_path}")
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create all sheets
    create_network_sheet(wb)
    create_producers_sheet(wb)
    create_storage_sheet(wb)
    create_pipes_sheet(wb)
    create_consumers_sheet(wb)
    create_timeseries_sheet(wb)
    create_instructions_sheet(wb)

    # Save
    wb.save(output_path)
    print(f"✓ Template created successfully!")
    print(f"\nNext steps:")
    print(f"1. Open {output_path} in Excel")
    print(f"2. Fill out the 6 data sheets")
    print(f"3. Convert to YAML:")
    print(f"   python -c \"from energis.utils.thermal_network_excel_parser import ThermalNetworkExcelParser; \\")
    print(f"              parser = ThermalNetworkExcelParser('{output_path}'); \\")
    print(f"              parser.save_yaml('my_network.scenario.yaml')\"")
    print(f"\nDocumentation:")
    print(f"- docs/excel_import_feature.md")
    print(f"- docs/brownfield_quickstart_guide.md")


if __name__ == '__main__':
    main()
