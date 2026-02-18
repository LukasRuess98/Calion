"""Model inspection and export utilities for Pyomo optimization models.

This module provides functionality to export Pyomo model structure (variables,
parameters, constraints, objectives) to Excel and Markdown formats for review
before solver execution.
"""

from __future__ import annotations

from typing import Any, Dict, List
import os
import json

from energis.logging_config import get_logger

logger = get_logger(__name__)

try:
    import pyomo.environ as pyo
    HAVE_PYOMO = True
except Exception:
    HAVE_PYOMO = False
    pyo = None

try:
    import matplotlib
    matplotlib.use('Agg')  # Non-interactive backend
    import matplotlib.pyplot as plt
    import numpy as np
    HAVE_MATPLOTLIB = True
except Exception:
    HAVE_MATPLOTLIB = False
    plt = None
    np = None


__all__ = [
    "export_model_structure",
    "inspect_pyomo_model",
    "create_model_plots",
]


def _safe_value(obj: Any, default: str = "N/A") -> Any:
    """Safely extract value from Pyomo object."""
    if not HAVE_PYOMO:
        return default

    try:
        if hasattr(obj, 'value'):
            val = pyo.value(obj)
            if val is None:
                return default
            return val
        return obj
    except Exception:
        return default


def _safe_bounds(var: Any) -> tuple[Any, Any]:
    """Safely extract bounds from Pyomo variable."""
    if not HAVE_PYOMO:
        return (None, None)

    try:
        lb = var.lb if hasattr(var, 'lb') else None
        ub = var.ub if hasattr(var, 'ub') else None
        return (lb, ub)
    except Exception:
        return (None, None)


def _format_constraint_expression(constraint: Any, max_length: int = 200) -> str:
    """Format constraint expression for display."""
    if not HAVE_PYOMO:
        return "Pyomo not available"

    try:
        expr_str = str(constraint.expr)
        if len(expr_str) > max_length:
            return expr_str[:max_length-3] + "..."
        return expr_str
    except Exception:
        return "Error formatting expression"


def inspect_pyomo_model(model: Any) -> Dict[str, Any]:
    """Inspect Pyomo model and extract structure information.

    Args:
        model: Pyomo ConcreteModel or AbstractModel

    Returns:
        Dictionary with model structure:
            - parameters: List of parameter dictionaries
            - variables: List of variable dictionaries
            - constraints: List of constraint dictionaries
            - objectives: List of objective dictionaries
            - sets: List of set dictionaries
            - summary: Model summary statistics
    """
    if not HAVE_PYOMO or model is None:
        return {
            "parameters": [],
            "variables": [],
            "constraints": [],
            "objectives": [],
            "sets": [],
            "summary": {"error": "Pyomo not available or model is None"}
        }

    inspection = {
        "parameters": [],
        "variables": [],
        "constraints": [],
        "objectives": [],
        "sets": [],
        "summary": {}
    }

    # Extract Sets
    for component in model.component_objects(pyo.Set, active=True):
        def _is_finite(c):
            # Pyomo renamed is_finite() to isfinite() in newer versions
            if hasattr(c, "isfinite"):
                return c.isfinite()
            return c.is_finite()

        set_info = {
            "name": str(component.name),
            "size": len(component) if _is_finite(component) else "infinite",
            "type": "Set",
        }

        # For small sets, show elements
        if _is_finite(component) and len(component) <= 20:
            set_info["elements"] = list(component)

        inspection["sets"].append(set_info)

    # Extract Parameters
    for component in model.component_objects(pyo.Param, active=True):
        param_info = {
            "name": str(component.name),
            "type": "Parameter",
            "indexed": component.is_indexed(),
            "size": len(component) if component.is_indexed() else 1,
        }

        # Get domain info
        try:
            if hasattr(component, 'domain'):
                param_info["domain"] = str(component.domain)
        except Exception:
            pass

        # For scalar or small indexed params, get values
        if not component.is_indexed():
            param_info["value"] = _safe_value(component)
        elif len(component) <= 10:
            param_info["values"] = {str(k): _safe_value(component[k]) for k in component}
        else:
            # Sample first few values
            keys = list(component.keys())[:5]
            param_info["sample_values"] = {str(k): _safe_value(component[k]) for k in keys}
            param_info["note"] = f"Showing 5 of {len(component)} values"

        inspection["parameters"].append(param_info)

    # Extract Variables
    for component in model.component_objects(pyo.Var, active=True):
        var_info = {
            "name": str(component.name),
            "type": "Variable",
            "indexed": component.is_indexed(),
            "size": len(component) if component.is_indexed() else 1,
        }

        # Get domain info
        try:
            if hasattr(component, 'domain'):
                domain_str = str(component.domain)
                if 'NonNegativeReals' in domain_str:
                    var_info["domain"] = "NonNegativeReals"
                elif 'Binary' in domain_str:
                    var_info["domain"] = "Binary"
                elif 'Reals' in domain_str:
                    var_info["domain"] = "Reals"
                else:
                    var_info["domain"] = domain_str
        except Exception:
            pass

        # Get bounds for scalar or sample for indexed
        if not component.is_indexed():
            lb, ub = _safe_bounds(component)
            var_info["lower_bound"] = lb
            var_info["upper_bound"] = ub
        elif len(component) <= 10:
            var_info["bounds"] = {str(k): _safe_bounds(component[k]) for k in component}
        else:
            # Sample first few bounds
            keys = list(component.keys())[:5]
            var_info["sample_bounds"] = {str(k): _safe_bounds(component[k]) for k in keys}
            var_info["note"] = f"Showing 5 of {len(component)} variables"

        inspection["variables"].append(var_info)

    # Extract Constraints
    for component in model.component_objects(pyo.Constraint, active=True):
        constr_info = {
            "name": str(component.name),
            "type": "Constraint",
            "indexed": component.is_indexed(),
            "size": len(component) if component.is_indexed() else 1,
        }

        # Get expression for scalar or sample for indexed
        if not component.is_indexed():
            constr_info["expression"] = _format_constraint_expression(component)
        elif len(component) <= 5:
            constr_info["expressions"] = {
                str(k): _format_constraint_expression(component[k])
                for k in component
            }
        else:
            # Sample first few expressions
            keys = list(component.keys())[:3]
            constr_info["sample_expressions"] = {
                str(k): _format_constraint_expression(component[k])
                for k in keys
            }
            constr_info["note"] = f"Showing 3 of {len(component)} constraints"

        inspection["constraints"].append(constr_info)

    # Extract Objectives
    for component in model.component_objects(pyo.Objective, active=True):
        obj_info = {
            "name": str(component.name),
            "type": "Objective",
            "sense": str(component.sense) if hasattr(component, 'sense') else "minimize",
            "indexed": component.is_indexed(),
        }

        # Get expression
        if not component.is_indexed():
            obj_info["expression"] = _format_constraint_expression(component)

        inspection["objectives"].append(obj_info)

    # Summary statistics
    inspection["summary"] = {
        "model_name": str(model.name) if hasattr(model, 'name') else "Unknown",
        "num_sets": len(inspection["sets"]),
        "num_parameters": len(inspection["parameters"]),
        "num_variables": sum(p["size"] for p in inspection["variables"]),
        "num_constraints": sum(c["size"] for c in inspection["constraints"]),
        "num_objectives": len(inspection["objectives"]),
        "variable_types": len(inspection["variables"]),
        "constraint_types": len(inspection["constraints"]),
    }

    return inspection


def _write_markdown_report(inspection: Dict[str, Any], filepath: str) -> None:
    """Write model inspection to Markdown file."""

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write("# Pyomo Model Structure Export\n\n")

        # Summary
        f.write("## Summary\n\n")
        summary = inspection.get("summary", {})
        f.write(f"- **Model Name**: {summary.get('model_name', 'N/A')}\n")
        f.write(f"- **Sets**: {summary.get('num_sets', 0)}\n")
        f.write(f"- **Parameters**: {summary.get('num_parameters', 0)}\n")
        f.write(f"- **Variables**: {summary.get('num_variables', 0)} ({summary.get('variable_types', 0)} types)\n")
        f.write(f"- **Constraints**: {summary.get('num_constraints', 0)} ({summary.get('constraint_types', 0)} types)\n")
        f.write(f"- **Objectives**: {summary.get('num_objectives', 0)}\n\n")

        # Sets
        if inspection.get("sets"):
            f.write("## Sets\n\n")
            f.write("| Name | Size | Elements |\n")
            f.write("|------|------|----------|\n")
            for s in inspection["sets"]:
                elements = s.get("elements", "...")
                if isinstance(elements, list):
                    elements = ", ".join(str(e) for e in elements[:10])
                f.write(f"| {s['name']} | {s.get('size', 'N/A')} | {elements} |\n")
            f.write("\n")

        # Parameters
        if inspection.get("parameters"):
            f.write("## Parameters\n\n")
            f.write("| Name | Type | Size | Domain | Value/Sample |\n")
            f.write("|------|------|------|--------|-------------|\n")
            for p in inspection["parameters"]:
                value_str = ""
                if "value" in p:
                    value_str = str(p["value"])
                elif "sample_values" in p:
                    samples = p["sample_values"]
                    value_str = ", ".join(f"{k}={v}" for k, v in list(samples.items())[:3])
                    if len(samples) > 3:
                        value_str += ", ..."

                f.write(f"| {p['name']} | {p.get('type', 'N/A')} | {p.get('size', 1)} | "
                       f"{p.get('domain', 'N/A')} | {value_str} |\n")
            f.write("\n")

        # Variables
        if inspection.get("variables"):
            f.write("## Variables\n\n")
            f.write("| Name | Type | Size | Domain | Bounds |\n")
            f.write("|------|------|------|--------|--------|\n")
            for v in inspection["variables"]:
                bounds_str = ""
                if "lower_bound" in v and "upper_bound" in v:
                    lb = v["lower_bound"] if v["lower_bound"] is not None else "-∞"
                    ub = v["upper_bound"] if v["upper_bound"] is not None else "+∞"
                    bounds_str = f"[{lb}, {ub}]"
                elif "sample_bounds" in v:
                    samples = v["sample_bounds"]
                    first_key = list(samples.keys())[0]
                    lb, ub = samples[first_key]
                    lb = lb if lb is not None else "-∞"
                    ub = ub if ub is not None else "+∞"
                    bounds_str = f"Example: [{lb}, {ub}]"

                f.write(f"| {v['name']} | {v.get('type', 'N/A')} | {v.get('size', 1)} | "
                       f"{v.get('domain', 'N/A')} | {bounds_str} |\n")
            f.write("\n")

        # Constraints
        if inspection.get("constraints"):
            f.write("## Constraints\n\n")
            for c in inspection["constraints"]:
                f.write(f"### {c['name']}\n\n")
                f.write(f"- **Type**: {c.get('type', 'N/A')}\n")
                f.write(f"- **Size**: {c.get('size', 1)}\n")

                if "expression" in c:
                    f.write(f"- **Expression**: `{c['expression']}`\n")
                elif "sample_expressions" in c:
                    f.write("- **Sample Expressions**:\n")
                    for k, expr in c["sample_expressions"].items():
                        f.write(f"  - `{k}`: {expr}\n")
                    if c.get("note"):
                        f.write(f"  - *{c['note']}*\n")
                f.write("\n")

        # Objectives
        if inspection.get("objectives"):
            f.write("## Objectives\n\n")
            for o in inspection["objectives"]:
                f.write(f"### {o['name']}\n\n")
                f.write(f"- **Sense**: {o.get('sense', 'minimize')}\n")
                if "expression" in o:
                    f.write(f"- **Expression**: `{o['expression']}`\n")
                f.write("\n")


def _write_excel_report(inspection: Dict[str, Any], filepath: str) -> None:
    """Write model inspection to Excel file."""

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to simple xlsx if openpyxl not available
        from energis.io.exporter import _write_simple_xlsx

        sheets = {}

        # Summary sheet
        summary_rows = [["Metric", "Value"]]
        for k, v in inspection.get("summary", {}).items():
            summary_rows.append([k, v])
        sheets["Summary"] = summary_rows

        # Parameters sheet
        param_rows = [["Name", "Type", "Size", "Domain", "Value/Sample"]]
        for p in inspection.get("parameters", []):
            value_str = str(p.get("value", p.get("sample_values", p.get("note", ""))))
            param_rows.append([
                p.get("name", ""),
                p.get("type", ""),
                p.get("size", ""),
                p.get("domain", ""),
                value_str
            ])
        sheets["Parameters"] = param_rows

        # Variables sheet
        var_rows = [["Name", "Type", "Size", "Domain", "Bounds"]]
        for v in inspection.get("variables", []):
            bounds_str = ""
            if "lower_bound" in v and "upper_bound" in v:
                bounds_str = f"[{v['lower_bound']}, {v['upper_bound']}]"
            elif "sample_bounds" in v:
                bounds_str = str(v["sample_bounds"])

            var_rows.append([
                v.get("name", ""),
                v.get("type", ""),
                v.get("size", ""),
                v.get("domain", ""),
                bounds_str
            ])
        sheets["Variables"] = var_rows

        # Constraints sheet
        constr_rows = [["Name", "Type", "Size", "Expression/Note"]]
        for c in inspection.get("constraints", []):
            expr_str = c.get("expression", c.get("note", ""))
            if "sample_expressions" in c:
                expr_str = str(c["sample_expressions"])

            constr_rows.append([
                c.get("name", ""),
                c.get("type", ""),
                c.get("size", ""),
                expr_str
            ])
        sheets["Constraints"] = constr_rows

        # Objectives sheet
        obj_rows = [["Name", "Sense", "Expression"]]
        for o in inspection.get("objectives", []):
            obj_rows.append([
                o.get("name", ""),
                o.get("sense", ""),
                o.get("expression", "")
            ])
        sheets["Objectives"] = obj_rows

        _write_simple_xlsx(filepath, sheets)
        return

    # Use openpyxl for better formatting
    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Metric", "Value"])
    for k, v in inspection.get("summary", {}).items():
        ws_summary.append([k, v])

    # Parameters sheet
    ws_params = wb.create_sheet("Parameters")
    ws_params.append(["Name", "Type", "Size", "Domain", "Value/Sample"])
    for p in inspection.get("parameters", []):
        value_str = str(p.get("value", p.get("sample_values", p.get("note", ""))))
        ws_params.append([
            p.get("name", ""),
            p.get("type", ""),
            p.get("size", ""),
            p.get("domain", ""),
            value_str
        ])

    # Variables sheet
    ws_vars = wb.create_sheet("Variables")
    ws_vars.append(["Name", "Type", "Size", "Domain", "Lower Bound", "Upper Bound", "Note"])
    for v in inspection.get("variables", []):
        lb = v.get("lower_bound", "")
        ub = v.get("upper_bound", "")
        note = v.get("note", "")

        if "sample_bounds" in v:
            samples = v["sample_bounds"]
            first_key = list(samples.keys())[0] if samples else ""
            if first_key:
                lb, ub = samples[first_key]
            note = v.get("note", "")

        ws_vars.append([
            v.get("name", ""),
            v.get("type", ""),
            v.get("size", ""),
            v.get("domain", ""),
            lb if lb is not None else "",
            ub if ub is not None else "",
            note
        ])

    # Constraints sheet
    ws_constr = wb.create_sheet("Constraints")
    ws_constr.append(["Name", "Type", "Size", "Expression", "Note"])
    for c in inspection.get("constraints", []):
        expr_str = c.get("expression", "")
        note = c.get("note", "")

        if "sample_expressions" in c:
            # Show first sample expression
            samples = c["sample_expressions"]
            first_key = list(samples.keys())[0] if samples else ""
            if first_key:
                expr_str = f"{first_key}: {samples[first_key]}"

        ws_constr.append([
            c.get("name", ""),
            c.get("type", ""),
            c.get("size", ""),
            expr_str,
            note
        ])

    # Objectives sheet
    ws_obj = wb.create_sheet("Objectives")
    ws_obj.append(["Name", "Sense", "Expression"])
    for o in inspection.get("objectives", []):
        ws_obj.append([
            o.get("name", ""),
            o.get("sense", ""),
            o.get("expression", "")
        ])

    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filepath)


def create_model_plots(
    model: Any,
    inspection: Dict[str, Any],
    output_dir: str,
    prefix: str = "model"
) -> List[str]:
    """Create visualization plots for Pyomo model structure.

    Args:
        model: Pyomo ConcreteModel
        inspection: Model inspection dictionary from inspect_pyomo_model()
        output_dir: Directory for plot outputs
        prefix: Filename prefix for plots

    Returns:
        List of paths to created plot files
    """
    if not HAVE_MATPLOTLIB or not HAVE_PYOMO or model is None:
        return []

    os.makedirs(output_dir, exist_ok=True)
    plot_files = []

    try:
        # Set style for professional-looking plots
        plt.style.use('seaborn-v0_8-darkgrid')
    except Exception:
        try:
            plt.style.use('seaborn-darkgrid')
        except Exception:
            pass  # Use default style

    # Plot 1: Model Size Overview
    try:
        fig, ax = plt.subplots(figsize=(10, 6))
        summary = inspection.get("summary", {})

        categories = ['Sets', 'Parameters', 'Variables', 'Constraints', 'Objectives']
        values = [
            summary.get('num_sets', 0),
            summary.get('num_parameters', 0),
            summary.get('num_variables', 0),
            summary.get('num_constraints', 0),
            summary.get('num_objectives', 0)
        ]

        colors = ['#2ecc71', '#3498db', '#e74c3c', '#f39c12', '#9b59b6']
        bars = ax.bar(categories, values, color=colors, alpha=0.7, edgecolor='black', linewidth=1.5)

        # Add value labels on bars
        for bar, val in zip(bars, values):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{int(val):,}',
                   ha='center', va='bottom', fontsize=10, fontweight='bold')

        ax.set_ylabel('Count', fontsize=12, fontweight='bold')
        ax.set_title(f'Model Structure Overview: {summary.get("model_name", "Unknown")}',
                    fontsize=14, fontweight='bold', pad=20)
        ax.set_yscale('log')
        ax.grid(True, alpha=0.3, axis='y')

        plt.tight_layout()
        path = os.path.join(output_dir, f"{prefix}_01_overview.png")
        plt.savefig(path, dpi=300, bbox_inches='tight')
        plt.close()
        plot_files.append(path)
        logger.info(f"[PLOT] Created: {os.path.basename(path)}")
    except Exception as e:
        logger.info(f"[PLOT] Warning: Could not create overview plot: {e}")

    # Plot 2: Variable Types Distribution
    try:
        fig, ax = plt.subplots(figsize=(10, 6))

        var_types = {}
        for v in inspection.get("variables", []):
            domain = v.get("domain", "Unknown")
            size = v.get("size", 1)
            var_types[domain] = var_types.get(domain, 0) + size

        if var_types:
            domains = list(var_types.keys())
            counts = list(var_types.values())

            colors_map = {
                'NonNegativeReals': '#3498db',
                'Binary': '#e74c3c',
                'Reals': '#2ecc71',
                'PositiveReals': '#f39c12',
            }
            colors = [colors_map.get(d, '#95a5a6') for d in domains]

            wedges, texts, autotexts = ax.pie(counts, labels=domains, autopct='%1.1f%%',
                                               colors=colors, startangle=90,
                                               textprops={'fontsize': 10, 'fontweight': 'bold'})

            # Add count labels
            for i, (domain, count) in enumerate(zip(domains, counts)):
                texts[i].set_text(f'{domain}\n({count:,})')

            ax.set_title('Variable Types Distribution', fontsize=14, fontweight='bold', pad=20)

            plt.tight_layout()
            path = os.path.join(output_dir, f"{prefix}_02_variable_types.png")
            plt.savefig(path, dpi=300, bbox_inches='tight')
            plt.close()
            plot_files.append(path)
            logger.info(f"[PLOT] Created: {os.path.basename(path)}")
    except Exception as e:
        logger.info(f"[PLOT] Warning: Could not create variable types plot: {e}")

    # Plot 3: Constraint Sizes
    try:
        fig, ax = plt.subplots(figsize=(12, 6))

        constraints = inspection.get("constraints", [])[:20]  # Top 20
        if constraints:
            names = [c.get("name", "")[:30] for c in constraints]
            sizes = [c.get("size", 0) for c in constraints]

            bars = ax.barh(names, sizes, color='#e74c3c', alpha=0.7, edgecolor='black', linewidth=1.0)

            # Add value labels
            for bar, size in zip(bars, sizes):
                width = bar.get_width()
                ax.text(width, bar.get_y() + bar.get_height()/2.,
                       f' {int(size):,}',
                       ha='left', va='center', fontsize=9, fontweight='bold')

            ax.set_xlabel('Number of Constraints', fontsize=12, fontweight='bold')
            ax.set_title('Constraint Sizes (Top 20)', fontsize=14, fontweight='bold', pad=20)
            ax.set_xscale('log')
            ax.grid(True, alpha=0.3, axis='x')

            plt.tight_layout()
            path = os.path.join(output_dir, f"{prefix}_03_constraint_sizes.png")
            plt.savefig(path, dpi=300, bbox_inches='tight')
            plt.close()
            plot_files.append(path)
            logger.info(f"[PLOT] Created: {os.path.basename(path)}")
    except Exception as e:
        logger.info(f"[PLOT] Warning: Could not create constraint sizes plot: {e}")

    # Plot 4: Parameter Time Series (if available)
    try:
        # Look for common time series parameters
        timeseries_params = {}
        for p in inspection.get("parameters", []):
            name = p.get("name", "")
            if p.get("indexed", False) and p.get("size", 0) > 100:  # Likely time series
                if any(keyword in name.lower() for keyword in
                      ['preis', 'price', 'bedarf', 'demand', 'co2', 'temperatur', 'temp']):
                    if "sample_values" in p:
                        # Extract values
                        samples = p["sample_values"]
                        values = [v for v in samples.values() if isinstance(v, (int, float))]
                        if values:
                            timeseries_params[name] = values[:1000]  # First 1000 values

        if timeseries_params:
            n_plots = len(timeseries_params)
            fig, axes = plt.subplots(n_plots, 1, figsize=(14, 4*n_plots))

            if n_plots == 1:
                axes = [axes]

            colors_list = ['#3498db', '#e74c3c', '#2ecc71', '#f39c12', '#9b59b6']

            for i, (name, values) in enumerate(timeseries_params.items()):
                ax = axes[i]
                x = list(range(len(values)))
                color = colors_list[i % len(colors_list)]

                ax.plot(x, values, linewidth=2, color=color, alpha=0.8)
                ax.fill_between(x, values, alpha=0.3, color=color)
                ax.set_ylabel(name, fontsize=11, fontweight='bold')
                ax.grid(True, alpha=0.3)

                # Add statistics
                mean_val = np.mean(values)
                ax.axhline(mean_val, color='red', linestyle='--', linewidth=1.5, alpha=0.7, label=f'Mean: {mean_val:.2f}')
                ax.legend(loc='upper right', fontsize=9)

            axes[-1].set_xlabel('Time Step', fontsize=12, fontweight='bold')
            fig.suptitle('Key Parameter Time Series', fontsize=14, fontweight='bold', y=0.995)

            plt.tight_layout()
            path = os.path.join(output_dir, f"{prefix}_04_parameter_timeseries.png")
            plt.savefig(path, dpi=300, bbox_inches='tight')
            plt.close()
            plot_files.append(path)
            logger.info(f"[PLOT] Created: {os.path.basename(path)}")
    except Exception as e:
        logger.info(f"[PLOT] Warning: Could not create parameter time series plot: {e}")

    # Plot 5: Variable Bounds Overview
    try:
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))

        # Collect bounds info
        bounded_vars = []
        unbounded_vars = []

        for v in inspection.get("variables", [])[:30]:  # Top 30
            name = v.get("name", "")
            lb = v.get("lower_bound")
            ub = v.get("upper_bound")

            if lb is not None and ub is not None and lb != float('-inf') and ub != float('inf'):
                bounded_vars.append((name, lb, ub))
            else:
                unbounded_vars.append(name)

        # Left plot: Bounded variables
        if bounded_vars:
            names = [n[:25] for n, _, _ in bounded_vars[:15]]  # Top 15
            lbs = [lb if lb != float('-inf') else 0 for _, lb, _ in bounded_vars[:15]]
            ubs = [ub if ub != float('inf') else 1e6 for _, _, ub in bounded_vars[:15]]

            y_pos = np.arange(len(names))

            # Plot ranges as horizontal bars
            for i, (lb, ub) in enumerate(zip(lbs, ubs)):
                ax1.barh(i, ub - lb, left=lb, height=0.6,
                        color='#3498db', alpha=0.7, edgecolor='black')

            ax1.set_yticks(y_pos)
            ax1.set_yticklabels(names, fontsize=9)
            ax1.set_xlabel('Value Range', fontsize=11, fontweight='bold')
            ax1.set_title('Variable Bounds (Top 15)', fontsize=12, fontweight='bold')
            ax1.grid(True, alpha=0.3, axis='x')

        # Right plot: Bounded vs Unbounded
        bounded_count = len(bounded_vars)
        unbounded_count = len(unbounded_vars)

        if bounded_count > 0 or unbounded_count > 0:
            labels = ['Bounded', 'Unbounded']
            sizes = [bounded_count, unbounded_count]
            colors = ['#2ecc71', '#e74c3c']
            explode = (0.05, 0)

            wedges, texts, autotexts = ax2.pie(sizes, explode=explode, labels=labels,
                                               autopct='%1.1f%%', colors=colors,
                                               startangle=90, textprops={'fontweight': 'bold'})

            ax2.set_title('Variables: Bounded vs Unbounded', fontsize=12, fontweight='bold')

        plt.tight_layout()
        path = os.path.join(output_dir, f"{prefix}_05_variable_bounds.png")
        plt.savefig(path, dpi=300, bbox_inches='tight')
        plt.close()
        plot_files.append(path)
        logger.info(f"[PLOT] Created: {os.path.basename(path)}")
    except Exception as e:
        logger.info(f"[PLOT] Warning: Could not create variable bounds plot: {e}")

    # Plot 6: Model Complexity Heatmap
    try:
        fig, ax = plt.subplots(figsize=(10, 8))

        # Create complexity matrix
        components = ['Sets', 'Parameters', 'Variables', 'Constraints']
        metrics = ['Count', 'Indexed', 'Scalar', 'Size (Total)']

        summary = inspection.get("summary", {})

        # Build data matrix
        data = np.zeros((len(components), len(metrics)))

        # Sets
        data[0, 0] = summary.get('num_sets', 0)

        # Parameters
        params = inspection.get("parameters", [])
        data[1, 0] = len(params)
        data[1, 1] = sum(1 for p in params if p.get("indexed", False))
        data[1, 2] = sum(1 for p in params if not p.get("indexed", False))
        data[1, 3] = sum(p.get("size", 1) for p in params)

        # Variables
        vars_list = inspection.get("variables", [])
        data[2, 0] = len(vars_list)
        data[2, 1] = sum(1 for v in vars_list if v.get("indexed", False))
        data[2, 2] = sum(1 for v in vars_list if not v.get("indexed", False))
        data[2, 3] = summary.get('num_variables', 0)

        # Constraints
        constrs = inspection.get("constraints", [])
        data[3, 0] = len(constrs)
        data[3, 1] = sum(1 for c in constrs if c.get("indexed", False))
        data[3, 2] = sum(1 for c in constrs if not c.get("indexed", False))
        data[3, 3] = summary.get('num_constraints', 0)

        # Normalize for better visualization (log scale)
        data_log = np.log10(data + 1)  # +1 to avoid log(0)

        im = ax.imshow(data_log, cmap='YlOrRd', aspect='auto')

        # Set ticks and labels
        ax.set_xticks(np.arange(len(metrics)))
        ax.set_yticks(np.arange(len(components)))
        ax.set_xticklabels(metrics, fontsize=10, fontweight='bold')
        ax.set_yticklabels(components, fontsize=11, fontweight='bold')

        # Rotate x labels
        plt.setp(ax.get_xticklabels(), rotation=45, ha="right", rotation_mode="anchor")

        # Add text annotations
        for i in range(len(components)):
            for j in range(len(metrics)):
                val = int(data[i, j])
                if val > 0:
                    ax.text(j, i, f'{val:,}',
                           ha="center", va="center", color="black",
                           fontsize=9, fontweight='bold')

        ax.set_title('Model Complexity Matrix', fontsize=14, fontweight='bold', pad=20)

        # Add colorbar
        cbar = plt.colorbar(im, ax=ax)
        cbar.set_label('log10(Count + 1)', rotation=270, labelpad=20, fontweight='bold')

        plt.tight_layout()
        path = os.path.join(output_dir, f"{prefix}_06_complexity_matrix.png")
        plt.savefig(path, dpi=300, bbox_inches='tight')
        plt.close()
        plot_files.append(path)
        logger.info(f"[PLOT] Created: {os.path.basename(path)}")
    except Exception as e:
        logger.info(f"[PLOT] Warning: Could not create complexity matrix plot: {e}")

    return plot_files


def export_model_structure(
    model: Any,
    output_dir: str,
    prefix: str = "model_structure"
) -> Dict[str, str]:
    """Export Pyomo model structure to Excel and Markdown files.

    Args:
        model: Pyomo ConcreteModel to export
        output_dir: Directory for output files
        prefix: Filename prefix for exports

    Returns:
        Dictionary with paths to created files:
            - excel_path: Path to Excel export
            - markdown_path: Path to Markdown export
            - json_path: Path to JSON export
    """
    os.makedirs(output_dir, exist_ok=True)

    # Inspect model
    logger.info(f"[MODEL_EXPORT] Inspecting Pyomo model...")
    inspection = inspect_pyomo_model(model)

    # Create output paths
    excel_path = os.path.join(output_dir, f"{prefix}.xlsx")
    markdown_path = os.path.join(output_dir, f"{prefix}.md")
    json_path = os.path.join(output_dir, f"{prefix}.json")

    # Export to Excel
    logger.info(f"[MODEL_EXPORT] Writing Excel report: {excel_path}")
    _write_excel_report(inspection, excel_path)

    # Export to Markdown
    logger.info(f"[MODEL_EXPORT] Writing Markdown report: {markdown_path}")
    _write_markdown_report(inspection, markdown_path)

    # Export to JSON (full inspection data)
    logger.info(f"[MODEL_EXPORT] Writing JSON report: {json_path}")
    # Make JSON serializable
    json_safe_inspection = json.loads(json.dumps(inspection, default=str))
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(json_safe_inspection, f, indent=2)

    summary = inspection.get("summary", {})
    logger.info(f"[MODEL_EXPORT] Model exported successfully:")
    logger.info(f"  - Variables: {summary.get('num_variables', 0)}")
    logger.info(f"  - Constraints: {summary.get('num_constraints', 0)}")
    logger.info(f"  - Parameters: {summary.get('num_parameters', 0)}")
    logger.info(f"  - Objectives: {summary.get('num_objectives', 0)}")

    # Create visualization plots
    plot_paths = []
    try:
        logger.info(f"[MODEL_EXPORT] Creating visualization plots...")
        plot_paths = create_model_plots(model, inspection, output_dir, prefix=prefix)
        if plot_paths:
            logger.info(f"[MODEL_EXPORT] Created {len(plot_paths)} visualization plots")
    except Exception as exc:
        logger.info(f"[MODEL_EXPORT] Warning: Could not create plots: {exc}")

    return {
        "excel_path": excel_path,
        "markdown_path": markdown_path,
        "json_path": json_path,
        "plot_paths": plot_paths,
    }
