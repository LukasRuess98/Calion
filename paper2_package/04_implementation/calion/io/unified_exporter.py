"""
Unified Export Manager for Heat Planning Framework.

This module provides a single entry point for all export operations,
consolidating:
- Framework exports (Excel, CSV, JSON) — consolidated from legacy exporter.py
- Publication exports (LaTeX tables, KPIs)
- Plot exports (publication-ready figures)
- Dashboard exports
- Sensitivity analysis exports

Usage:
    >>> from calion.io.unified_exporter import UnifiedExporter
    >>> exporter = UnifiedExporter(workflow_result)
    >>> exporter.export_all("exports/my_scenario")

    # Or step by step:
    >>> exporter.export_data()        # CSV, Excel, JSON
    >>> exporter.export_publication() # LaTeX tables
    >>> exporter.export_plots()       # Figures
"""

from __future__ import annotations

import csv
import json
import math
import os
import time
from collections import OrderedDict
from collections.abc import Iterable, Mapping, MutableMapping, Sequence
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from numbers import Number
from typing import TYPE_CHECKING, Any
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd

from calion.logging_config import get_logger
from calion.utils.timeseries import TimeSeriesTable

logger = get_logger(__name__)

if TYPE_CHECKING:
    from calion.run.rolling_horizon import WorkflowResult

# ─ Declare optional dependencies ─
try:  # pragma: no cover - optional dependency
    from openpyxl import Workbook
    HAVE_OPENPYXL = True
except Exception:  # pragma: no cover
    Workbook = None
    HAVE_OPENPYXL = False

_XML_ESCAPE = {"'": "&apos;"}

from calion.io.exporters.multi_network_exporter import (  # noqa: F401
    export_multi_network_data,
    export_multi_network_latex,
    export_multi_network_plots,
    export_multi_network_results,
)

__all__ = [
    "HAVE_OPENPYXL",
    "export_multi_network_data",
    "export_multi_network_latex",
    "export_multi_network_plots",
    "export_multi_network_results",
    "export_scenario_bundle",
    "export_workflow",
    "export_with_sensitivity",
    "write_excel_workbook",
    "write_scenario_workbook",
    "write_timeseries_csv",
    "ExportConfig",
    "ExportResult",
    "UnifiedExporter",
]

# ═════════════════════════════════════════════════════════════════════════════
# LEGACY EXPORTER.PY UTILITIES (consolidated into unified_exporter)
# ═════════════════════════════════════════════════════════════════════════════


def _ensure_lengths(data: Mapping[str, Sequence[float]], expected: int) -> None:
    """Validate that all series have expected length."""
    for key, values in data.items():
        if len(values) != expected:
            raise ValueError(f"Series {key!r} has length {len(values)} instead of {expected}")


def write_timeseries_csv(
    path: str,
    table: TimeSeriesTable,
    extra: Mapping[str, Sequence[float]],
    *,
    decimal_separator: str = ",",
    alternate_path: str | None = None,
    alternate_decimal_separator: str = ".",
) -> None:
    """Write input and result time series to a semicolon separated CSV file.

    Parameters
    ----------
    path:
        Destination of the main CSV export.
    table:
        The base time series table to be exported.
    extra:
        Additional series that should be exported next to the main table.
    decimal_separator:
        Decimal separator used for *path*. Defaults to a comma which is
        convenient for German Excel installations.
    alternate_path:
        Optional second file that mirrors the exported content but allows
        specifying a different decimal separator. This can be helpful when the
        data should be double-checked in tools that prefer the conventional dot
        separator.
    alternate_decimal_separator:
        Decimal separator used for *alternate_path*. Ignored when no alternate
        path is supplied.
    """
    n = len(table)
    _ensure_lengths(extra, n)

    columns = ["timestamp", *table.columns, *list(extra.keys())]
    _write_csv(
        path,
        columns,
        table,
        extra,
        decimal_separator=decimal_separator,
    )

    if alternate_path:
        _write_csv(
            alternate_path,
            columns,
            table,
            extra,
            decimal_separator=alternate_decimal_separator,
        )


def _fmt_value(value: object, *, decimal_separator: str = ",") -> str:
    """Format value for CSV output."""
    if value is None:
        return ""

    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return ""
        try:
            decimal_value = _to_decimal(stripped)
        except InvalidOperation:
            return value
        text = _decimal_to_text(decimal_value)
        return _apply_decimal_separator(text, decimal_separator)

    if isinstance(value, Number) and not isinstance(value, bool):
        decimal_value = _to_decimal(value)
        text = _decimal_to_text(decimal_value)
        return _apply_decimal_separator(text, decimal_separator)

    return str(value)


def _write_csv(
    path: str,
    columns: Sequence[str],
    table: TimeSeriesTable,
    extra: Mapping[str, Sequence[float]],
    *,
    decimal_separator: str,
) -> None:
    """Write CSV file with given columns and data."""
    with open(path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(
            handle,
            delimiter=";",
            quoting=csv.QUOTE_MINIMAL,
            lineterminator="\n",
        )
        writer.writerow(columns)
        for idx, ts in enumerate(table.index):
            base = [ts.isoformat() if isinstance(ts, datetime) else str(ts)]
            base.extend(
                _fmt_value(table[col][idx], decimal_separator=decimal_separator)
                for col in table.columns
            )
            base.extend(
                _fmt_value(extra[name][idx], decimal_separator=decimal_separator)
                for name in extra
            )
            writer.writerow(base)


def _to_decimal(value: Number | Decimal | str) -> Decimal:
    """Convert value to Decimal."""
    if isinstance(value, Decimal):
        return value
    if isinstance(value, str):
        normalized = value.replace(" ", "").replace(",", ".")
        return Decimal(normalized)
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return Decimal(value)


def _decimal_to_text(value: Decimal) -> str:
    """Convert Decimal to text representation."""
    text = format(value.normalize(), "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def _apply_decimal_separator(text: str, decimal_separator: str) -> str:
    """Apply decimal separator to text."""
    if decimal_separator not in {",", "."}:
        raise ValueError(
            "decimal_separator must be ',' or '.' to avoid ambiguous CSV numbers"
        )
    if decimal_separator == ".":
        return text
    return text.replace(".", ",")


def _normalize_excel_value(value: object) -> object:
    """Convert arbitrary values into a representation accepted by openpyxl."""
    if value is None:
        return None
    if isinstance(value, (str, Number, bool)):
        return value
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    iso_formatter = getattr(value, "isoformat", None)
    if callable(iso_formatter):
        try:
            return iso_formatter()
        except Exception:
            pass
    return str(value)


def _append_table(ws: Any, headers: Sequence[str], rows: Iterable[Sequence[object]]) -> None:
    """Append table to Excel worksheet."""
    ws.append([_normalize_excel_value(header) for header in headers])
    for row in rows:
        ws.append([_normalize_excel_value(value) for value in row])


def _excel_safe_value(value: object) -> object:
    """Convert value to Excel-safe representation."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        value = float(value)
    if isinstance(value, Number) and not isinstance(value, bool):
        if isinstance(value, float) and not math.isfinite(value):
            return ""
        return value
    if isinstance(value, bool):
        return value
    if isinstance(value, (list, tuple, dict)):
        try:
            return json.dumps(value, ensure_ascii=False)
        except TypeError:
            return str(value)
    return str(value)


def _excel_safe_df(columns: Sequence[str], rows: Iterable[Mapping[str, object]]) -> list[list[object]]:
    """Convert dataframe rows to Excel-safe format."""
    safe_rows: list[list[object]] = []
    for row in rows:
        safe_row: list[object] = []
        for col in columns:
            safe_row.append(_excel_safe_value(row.get(col, "")))
        safe_rows.append(safe_row)
    return safe_rows


def _column_letter(idx: int) -> str:
    """Convert column index to Excel column letter."""
    if idx <= 0:
        return "A"
    result = ""
    while idx:
        idx, remainder = divmod(idx - 1, 26)
        result = chr(65 + remainder) + result
    return result or "A"


def _sheet_xml(rows: Sequence[Sequence[object]]) -> str:
    """Generate XML for worksheet."""
    parts = [
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>",
        "<worksheet xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\">",
        "<sheetData>",
    ]
    for row_idx, row in enumerate(rows, start=1):
        cells: list[str] = []
        for col_idx, value in enumerate(row, start=1):
            if value is None or value == "":
                continue
            text = str(value)
            text = text.replace("\r\n", "\n").replace("\r", "\n")
            text = escape(text, _XML_ESCAPE)
            text = text.replace("\n", "&#10;")
            cell_ref = f"{_column_letter(col_idx)}{row_idx}"
            cells.append(
                f"<c r=\"{cell_ref}\" t=\"inlineStr\"><is><t>{text}</t></is></c>"
            )
        if cells:
            parts.append(f"<row r=\"{row_idx}\">{''.join(cells)}</row>")
        else:
            parts.append(f"<row r=\"{row_idx}\"/>")
    parts.append("</sheetData></worksheet>")
    return "".join(parts)


def _workbook_xml(sheet_names: Sequence[str]) -> str:
    """Generate XML for workbook."""
    parts = [
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>",
        "<workbook xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\" xmlns:r=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships\">",
        "<sheets>",
    ]
    for idx, name in enumerate(sheet_names, start=1):
        safe_name = escape(str(name), _XML_ESCAPE)
        parts.append(
            f"<sheet name=\"{safe_name}\" sheetId=\"{idx}\" r:id=\"rId{idx}\"/>"
        )
    parts.append("</sheets></workbook>")
    return "".join(parts)


def _workbook_rels_xml(num_sheets: int) -> str:
    """Generate XML for workbook relationships."""
    parts = [
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>",
        "<Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\">",
    ]
    for idx in range(1, num_sheets + 1):
        parts.append(
            f"<Relationship Id=\"rId{idx}\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet\" Target=\"worksheets/sheet{idx}.xml\"/>"
        )
    parts.append("</Relationships>")
    return "".join(parts)


def _root_rels_xml() -> str:
    """Generate root relationships XML."""
    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>"
        "<Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\">"
        "<Relationship Id=\"rId1\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument\" Target=\"xl/workbook.xml\"/>"
        "<Relationship Id=\"rId2\" Type=\"http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties\" Target=\"docProps/core.xml\"/>"
        "<Relationship Id=\"rId3\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties\" Target=\"docProps/app.xml\"/>"
        "</Relationships>"
    )


def _content_types_xml(num_sheets: int) -> str:
    """Generate content types XML."""
    parts = [
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>",
        "<Types xmlns=\"http://schemas.openxmlformats.org/package/2006/content-types\">",
        "<Default Extension=\"rels\" ContentType=\"application/vnd.openxmlformats-package.relationships+xml\"/>",
        "<Default Extension=\"xml\" ContentType=\"application/xml\"/>",
        "<Override PartName=\"/xl/workbook.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml\"/>",
        "<Override PartName=\"/xl/styles.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml\"/>",
        "<Override PartName=\"/docProps/core.xml\" ContentType=\"application/vnd.openxmlformats-package.core-properties+xml\"/>",
        "<Override PartName=\"/docProps/app.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.extended-properties+xml\"/>",
    ]
    for idx in range(1, num_sheets + 1):
        parts.append(
            f"<Override PartName=\"/xl/worksheets/sheet{idx}.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml\"/>"
        )
    parts.append("</Types>")
    return "".join(parts)


def _styles_xml() -> str:
    """Generate styles XML."""
    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>"
        "<styleSheet xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\">"
        "<fonts count=\"1\"><font><sz val=\"11\"/><name val=\"Calibri\"/></font></fonts>"
        "<fills count=\"1\"><fill><patternFill patternType=\"none\"/></fill></fills>"
        "<borders count=\"1\"><border/></borders>"
        "<cellStyleXfs count=\"1\"><xf/></cellStyleXfs>"
        "<cellXfs count=\"1\"><xf xfId=\"0\"/></cellXfs>"
        "<cellStyles count=\"1\"><cellStyle name=\"Normal\" xfId=\"0\" builtinId=\"0\"/></cellStyles>"
        "</styleSheet>"
    )


def _app_xml(sheet_names: Sequence[str]) -> str:
    """Generate app.xml for metadata."""
    count = len(sheet_names)
    titles = "".join(
        f"<vt:lpstr>{escape(str(name), _XML_ESCAPE)}</vt:lpstr>" for name in sheet_names
    )
    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>"
        "<Properties xmlns=\"http://schemas.openxmlformats.org/officeDocument/2006/extended-properties\" xmlns:vt=\"http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes\">"
        "<Application>calion</Application>"
        "<HeadingPairs><vt:vector size=\"2\" baseType=\"variant\">"
        "<vt:variant><vt:lpstr>Worksheets</vt:lpstr></vt:variant>"
        f"<vt:variant><vt:i4>{count}</vt:i4></vt:variant>"
        "</vt:vector></HeadingPairs>"
        "<TitlesOfParts><vt:vector baseType=\"lpstr\" size=\"{count}\">"
        f"{titles}"
        "</vt:vector></TitlesOfParts>"
        "</Properties>"
    )


def _core_xml() -> str:
    """Generate core metadata XML."""
    from datetime import timezone
    now = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>"
        "<cp:coreProperties xmlns:cp=\"http://schemas.openxmlformats.org/package/2006/metadata/core-properties\" xmlns:dc=\"http://purl.org/dc/elements/1.1/\" xmlns:dcterms=\"http://purl.org/dc/terms/\" xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\">"
        "<dc:creator>calion</dc:creator>"
        "<cp:lastModifiedBy>calion</cp:lastModifiedBy>"
        f"<dcterms:created xsi:type=\"dcterms:W3CDTF\">{now}</dcterms:created>"
        f"<dcterms:modified xsi:type=\"dcterms:W3CDTF\">{now}</dcterms:modified>"
        "<cp:revision>1</cp:revision>"
        "</cp:coreProperties>"
    )


def _write_simple_xlsx(path: str, sheets: Mapping[str, Sequence[Sequence[object]]]) -> None:
    """Write simple XLSX file without openpyxl (minimal ZIP-based approach)."""
    directory = os.path.dirname(path)
    if directory:
        os.makedirs(directory, exist_ok=True)

    sheet_names = list(sheets.keys())
    with ZipFile(path, "w", ZIP_DEFLATED) as handle:
        handle.writestr("[Content_Types].xml", _content_types_xml(len(sheet_names)))
        handle.writestr("_rels/.rels", _root_rels_xml())
        handle.writestr("docProps/app.xml", _app_xml(sheet_names))
        handle.writestr("docProps/core.xml", _core_xml())
        handle.writestr("xl/workbook.xml", _workbook_xml(sheet_names))
        handle.writestr("xl/_rels/workbook.xml.rels", _workbook_rels_xml(len(sheet_names)))
        handle.writestr("xl/styles.xml", _styles_xml())
        for idx, (_name, rows) in enumerate(sheets.items(), start=1):
            handle.writestr(f"xl/worksheets/sheet{idx}.xml", _sheet_xml(rows))


def _build_meta_sheet(meta_sections: Mapping[str, Mapping[str, object]] | None) -> list[list[object]]:
    """Build metadata sheet."""
    if not meta_sections:
        return [["no metadata available"]]

    rows: list[list[object]] = []
    for section, entries in meta_sections.items():
        rows.append([_excel_safe_value(section)])
        if isinstance(entries, Mapping):
            for key, value in entries.items():
                rows.append([_excel_safe_value(key), _excel_safe_value(value)])
        else:
            rows.append(["value", _excel_safe_value(entries)])
        rows.append([])
    return rows


def _build_timeseries_sheet(
    sections: Sequence[Mapping[str, object]] | None,
) -> list[list[object]]:
    """Build timeseries sheet."""
    if not sections:
        return [["no timeseries available"]]

    rows: list[list[object]] = []
    for section in sections:
        label = _excel_safe_value(section.get("label", "timeseries"))
        timestamps = list(section.get("timestamps", []))
        series = section.get("series")
        rows.append([label])
        if not timestamps or not isinstance(series, Mapping) or not series:
            rows.append(["no data"])
            rows.append([])
            continue

        value_lists: OrderedDict[str, list[object]] = OrderedDict()
        for name, values in series.items():
            key = str(name)
            seq = list(values)
            if len(seq) < len(timestamps):
                seq.extend([""] * (len(timestamps) - len(seq)))
            elif len(seq) > len(timestamps):
                seq = seq[: len(timestamps)]
            value_lists[key] = seq

        headers = ["timestamp", *list(value_lists.keys())]
        rows.append([_excel_safe_value(header) for header in headers])

        data_rows = []
        for idx, ts in enumerate(timestamps):
            row: OrderedDict[str, object] = OrderedDict()
            row["timestamp"] = ts
            for name, seq in value_lists.items():
                row[name] = seq[idx]
            data_rows.append(row)

        rows.extend(_excel_safe_df(headers, data_rows))
        rows.append([])

    return rows or [["no timeseries available"]]


def _build_costs_sheet(costs: Mapping[str, Mapping[str, object]] | None) -> list[list[object]]:
    """Build costs sheet."""
    if not costs:
        return [["no costs available"]]

    rows: list[list[object]] = []
    for section, entries in costs.items():
        rows.append([_excel_safe_value(section)])
        if isinstance(entries, Mapping):
            for key, value in entries.items():
                rows.append([_excel_safe_value(key), _excel_safe_value(value)])
        else:
            rows.append(["value", _excel_safe_value(entries)])
        rows.append([])
    return rows


def _build_design_sheet(design: Mapping[str, object] | None) -> list[list[object]]:
    """Build design sheet."""
    if not design:
        return [["no design data available"]]

    rows: list[list[object]] = []
    heat_pumps = design.get("heat_pumps") if isinstance(design, Mapping) else None
    if isinstance(heat_pumps, Mapping) and heat_pumps:
        rows.append([_excel_safe_value("heat_pumps")])
        rows.append([_excel_safe_value("id"), _excel_safe_value("capacity_mw"), _excel_safe_value("build_binary")])
        for hp_id, metrics in heat_pumps.items():
            if isinstance(metrics, Mapping):
                rows.append([
                    _excel_safe_value(hp_id),
                    _excel_safe_value(metrics.get("capacity_mw")),
                    _excel_safe_value(metrics.get("build_binary")),
                ])
        rows.append([])

    storage = design.get("storage") if isinstance(design, Mapping) else None
    if isinstance(storage, Mapping) and storage:
        rows.append([_excel_safe_value("storage")])
        rows.append([
            _excel_safe_value("name"),
            _excel_safe_value("capacity_mwh"),
            _excel_safe_value("power_mw"),
            _excel_safe_value("build_binary"),
        ])
        rows.append([
            _excel_safe_value(storage.get("name")),
            _excel_safe_value(storage.get("capacity_mwh")),
            _excel_safe_value(storage.get("power_mw")),
            _excel_safe_value(storage.get("build_binary")),
        ])
        rows.append([])

    if not rows:
        rows.append(["no design data available"])
    return rows


def write_scenario_workbook(
    path: str,
    *,
    meta_sections: Mapping[str, Mapping[str, object]] | None = None,
    timeseries_sections: Sequence[Mapping[str, object]] | None = None,
    cost_sections: Mapping[str, Mapping[str, object]] | None = None,
    design: Mapping[str, object] | None = None,
) -> None:
    """Create scenario workbook with multiple sheets."""
    sheets: OrderedDict[str, list[list[object]]] = OrderedDict()
    sheets["Meta"] = _build_meta_sheet(meta_sections)
    sheets["Timeseries"] = _build_timeseries_sheet(timeseries_sections)
    sheets["Costs"] = _build_costs_sheet(cost_sections)
    sheets["Design"] = _build_design_sheet(design)
    _write_simple_xlsx(path, sheets)


def export_scenario_bundle(
    outdir: str,
    *,
    meta_sections: Mapping[str, Mapping[str, object]] | None = None,
    timeseries_sections: Sequence[Mapping[str, object]] | None = None,
    cost_sections: Mapping[str, Mapping[str, object]] | None = None,
    design: Mapping[str, object] | None = None,
    manifest: Mapping[str, object] | None = None,
) -> Mapping[str, object | None]:
    """Export scenario bundle (Excel workbook + design JSON + manifest)."""
    os.makedirs(outdir, exist_ok=True)

    workbook_path = os.path.join(outdir, "scenario.xlsx")
    write_scenario_workbook(
        workbook_path,
        meta_sections=meta_sections,
        timeseries_sections=timeseries_sections,
        cost_sections=cost_sections,
        design=design,
    )

    manifest_data: OrderedDict[str, object] = OrderedDict()
    if manifest:
        for key, value in manifest.items():
            manifest_data[key] = value

    files_entry: MutableMapping[str, object]
    raw_files = manifest_data.get("files")
    if isinstance(raw_files, MutableMapping):
        files_entry = raw_files
    else:
        files_entry = OrderedDict()
        manifest_data["files"] = files_entry

    files_entry["scenario_xlsx"] = os.path.basename(workbook_path)

    design_path: str | None = None
    if isinstance(design, Mapping) and (
        (isinstance(design.get("heat_pumps"), Mapping) and design.get("heat_pumps"))
        or (isinstance(design.get("storage"), Mapping) and design.get("storage"))
    ):
        design_path = os.path.join(outdir, "pf_design.json")
        with open(design_path, "w", encoding="utf-8") as handle:
            json.dump(design, handle, indent=2)
        files_entry["pf_design_json"] = os.path.basename(design_path)

    manifest_path = os.path.join(outdir, "manifest.json")
    with open(manifest_path, "w", encoding="utf-8") as handle:
        json.dump(manifest_data, handle, indent=2)

    return {
        "scenario_xlsx": workbook_path,
        "pf_design_json": design_path,
        "manifest_json": manifest_path,
    }


def write_excel_workbook(
    path: str,
    table: TimeSeriesTable,
    series: Mapping[str, Sequence[float]],
    summary_sections: Mapping[str, Mapping[str, object]],
    metadata_sections: Mapping[str, Mapping[str, object]] | None = None,
) -> None:
    """Create an Excel workbook containing inputs, results and summary tables."""
    if not HAVE_OPENPYXL:
        raise RuntimeError(
            "openpyxl ist nicht installiert – Excel-Export ist daher nicht möglich. Bitte openpyxl nachinstallieren."
        )

    n = len(table)
    _ensure_lengths(series, n)

    wb = Workbook()

    input_ws = wb.active
    input_ws.title = "input_timeseries"
    _append_table(
        input_ws,
        ["timestamp", *table.columns],
        (
            [table.index[i].isoformat() if isinstance(table.index[i], datetime) else str(table.index[i])] +
            [table[col][i] for col in table.columns]
            for i in range(n)
        ),
    )
    input_ws.freeze_panes = "B2"

    results_ws = wb.create_sheet("results_timeseries")
    series_headers = ["timestamp", *list(series.keys())]
    _append_table(
        results_ws,
        series_headers,
        (
            [table.index[i].isoformat() if isinstance(table.index[i], datetime) else str(table.index[i])] +
            [series[name][i] for name in series]
            for i in range(n)
        ),
    )
    results_ws.freeze_panes = "B2"

    summary_ws = wb.create_sheet("summary")
    for section, metrics in summary_sections.items():
        summary_ws.append([_normalize_excel_value(section)])
        for key, value in metrics.items():
            summary_ws.append([_normalize_excel_value(key), _normalize_excel_value(value)])
        summary_ws.append([])

    if metadata_sections:
        meta_ws = wb.create_sheet("metadata")
        for section, entries in metadata_sections.items():
            meta_ws.append([_normalize_excel_value(section)])
            for key, value in entries.items():
                if isinstance(value, (dict, list)):
                    normalized_value = json.dumps(value, ensure_ascii=False)
                else:
                    normalized_value = _normalize_excel_value(value)
                meta_ws.append([_normalize_excel_value(key), _normalize_excel_value(normalized_value)])
            meta_ws.append([])

    wb.save(path)


# ═════════════════════════════════════════════════════════════════════════════
# NEW UNIFIED EXPORTER CLASSES AND FUNCTIONS
# ═════════════════════════════════════════════════════════════════════════════


@dataclass
class ExportConfig:
    """Configuration for export operations."""

    # Data exports
    export_excel: bool = True
    export_csv: bool = True
    export_json: bool = True

    # Publication exports
    export_latex: bool = True
    latex_style: str = "booktabs"  # "booktabs", "applied_energies", "ieee"

    # Plot exports
    export_plots: bool = True
    plot_formats: list[str] = field(default_factory=lambda: ["png", "pdf"])
    plot_dpi: int = 300

    # Dashboard export
    export_dashboard: bool = False

    # Sensitivity (only if sensitivity results provided)
    export_sensitivity: bool = True

    # Multi-network exports (automatically enabled if multi-network results exist)
    export_multi_network: bool = True


@dataclass
class ExportResult:
    """Result of an export operation."""

    output_dir: str
    generated_files: dict[str, str] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)

    def summary(self) -> str:
        """Generate human-readable summary."""
        lines = [
            "=" * 60,
            "EXPORT SUMMARY",
            "=" * 60,
            f"Output directory: {self.output_dir}",
            f"Files generated: {len(self.generated_files)}",
        ]

        # Multi-network info
        if self.metadata.get("has_multi_network"):
            lines.append("Multi-network export: Yes")

        if self.errors:
            lines.append(f"Errors: {len(self.errors)}")
            for err in self.errors[:3]:
                lines.append(f"  - {err}")

        lines.append("")
        lines.append("Generated files:")

        # Group files by category
        categories = {
            "data": [],
            "publication": [],
            "figure": [],
            "multi_network": [],
            "dashboard": [],
            "other": [],
        }

        for category, path in sorted(self.generated_files.items()):
            if isinstance(path, (list, tuple)):
                path = path[0] if path else "unknown"
            basename = os.path.basename(str(path))

            if category.startswith("data_"):
                categories["data"].append((category, basename))
            elif category.startswith("latex_") or category.startswith("publication_"):
                categories["publication"].append((category, basename))
            elif category.startswith("figure_"):
                categories["figure"].append((category, basename))
            elif category.startswith("multi_network"):
                categories["multi_network"].append((category, basename))
            elif category.startswith("dashboard_"):
                categories["dashboard"].append((category, basename))
            else:
                categories["other"].append((category, basename))

        for cat_name, items in categories.items():
            if items:
                lines.append(f"  [{cat_name}]")
                for category, basename in items:
                    lines.append(f"    {category}: {basename}")

        lines.append("=" * 60)
        return "\n".join(lines)


class UnifiedExporter:
    """
    Unified export manager for workflow results.

    Consolidates all export operations into a single interface.

    Directory structure created:
        {output_dir}/
        ├── data/
        │   ├── timeseries.csv
        │   ├── summary.json
        │   └── scenario.xlsx
        ├── publication/
        │   ├── latex/
        │   │   ├── kpi_summary.tex
        │   │   └── cost_breakdown.tex
        │   └── kpi_summary.json
        ├── figures/
        │   ├── dispatch_profile.png
        │   ├── cost_breakdown.png
        │   └── ...
        └── dashboard/
            └── workflow_run.json
    """

    def __init__(
        self,
        workflow: WorkflowResult,
        config: ExportConfig | None = None,
    ):
        """
        Initialize the exporter.

        Parameters
        ----------
        workflow : WorkflowResult
            The workflow result from run_workflow()
        config : ExportConfig, optional
            Export configuration. Uses defaults if not provided.
        """
        self.workflow = workflow
        self.config = config or ExportConfig()

        # Determine the active result (MPC > RH > PF)
        self._active_result = (
            workflow.mpc_result or
            workflow.rh_result or
            workflow.pf_result
        )

        if self._active_result is None:
            raise ValueError("Workflow has no results (no PF, RH, or MPC result)")

        # Pre-compute common data structures
        self._summary_sections = self._build_summary_sections()
        self._timeseries_df = self._build_timeseries_df()

        # Multi-network data (if available)
        self._multi_network_results = self._extract_multi_network_results()

    def _build_summary_sections(self) -> dict[str, dict[str, Any]]:
        """Build unified summary sections from workflow results."""
        sections: dict[str, dict[str, Any]] = {}

        result = self._active_result

        # Copy existing summary sections
        if hasattr(result, "summary") and isinstance(result.summary, Mapping):
            for sec, vals in result.summary.items():
                sections[str(sec)] = dict(vals) if isinstance(vals, Mapping) else {"value": vals}

        # Build objective section from costs
        objective = sections.setdefault("objective", {})
        if hasattr(result, "costs") and isinstance(result.costs, Mapping):
            for key, value in result.costs.items():
                if isinstance(key, str) and key.startswith("objective."):
                    metric = key.split("objective.", 1)[1]
                    objective[metric] = value
                elif isinstance(key, str):
                    # Also include non-prefixed cost keys
                    objective[key] = value

        return sections

    def _build_timeseries_df(self) -> pd.DataFrame:
        """Build unified timeseries DataFrame."""
        result = self._active_result
        table = result.table

        # Start with input data
        df = pd.DataFrame(
            {col: table[col] for col in table.columns},
            index=table.index,
        )

        # Add result series
        if hasattr(result, "series"):
            for name, values in result.series.items():
                df[name] = values

        return df

    def _extract_multi_network_results(self) -> dict[str, Any] | None:
        """
        Extract multi-network results if available.

        Looks for multi-network data in:
        1. workflow.multi_network_results
        2. _active_result.multi_network
        3. Series with network-prefixed names
        """
        multi_results = None

        # Check workflow for multi-network results
        if hasattr(self.workflow, 'multi_network_results'):
            multi_results = self.workflow.multi_network_results

        # Check active result for multi-network data
        if multi_results is None and hasattr(self._active_result, 'multi_network'):
            multi_results = self._active_result.multi_network

        # Check summary sections for multi-network indicators
        if multi_results is None:
            for _section, data in self._summary_sections.items():
                if isinstance(data, dict):
                    if 'networks' in data or 'heat_exchangers' in data:
                        multi_results = data
                        break

        # Look for network-prefixed series (e.g., "hochtemp_T_supply")
        if multi_results is None and hasattr(self._active_result, 'series'):
            network_prefixes = set()
            for name in self._active_result.series.keys():
                if '_T_supply' in name or '_T_return' in name:
                    prefix = name.split('_T_')[0]
                    if prefix:
                        network_prefixes.add(prefix)

            if len(network_prefixes) > 1:
                # Multiple networks detected from series names
                multi_results = {
                    'networks': {p: {'id': p} for p in network_prefixes},
                    'detected_from_series': True,
                }

        return multi_results

    def has_multi_network_results(self) -> bool:
        """Check if multi-network results are available."""
        return self._multi_network_results is not None

    def export_all(
        self,
        output_dir: str | None = None,
        scenario_name: str | None = None,
    ) -> ExportResult:
        """
        Export all data in one call.

        Parameters
        ----------
        output_dir : str, optional
            Output directory. Auto-generated if not provided.
        scenario_name : str, optional
            Name for the scenario. Used in filenames.

        Returns
        -------
        ExportResult
            Summary of all generated files
        """
        # Create output directory
        if output_dir is None:
            output_dir = self._generate_output_dir(scenario_name)

        os.makedirs(output_dir, exist_ok=True)

        result = ExportResult(
            output_dir=output_dir,
            metadata={
                "timestamp": datetime.now().isoformat(),
                "scenario_name": scenario_name,
                "workflow_mode": self._get_workflow_mode(),
                "has_multi_network": self.has_multi_network_results(),
            }
        )

        logger.info("=" * 60)
        logger.info("UNIFIED EXPORT")
        logger.info(f"Output directory: {output_dir}")
        logger.info("=" * 60)

        # Export data (CSV, Excel, JSON)
        if self.config.export_csv or self.config.export_excel or self.config.export_json:
            try:
                data_files = self._export_data(output_dir)
                result.generated_files.update(data_files)
            except Exception as e:
                result.errors.append(f"Data export failed: {e}")
                logger.warning(f"Data export failed: {e}")

        # Export publication materials
        if self.config.export_latex:
            try:
                pub_files = self._export_publication(output_dir)
                result.generated_files.update(pub_files)
            except Exception as e:
                result.errors.append(f"Publication export failed: {e}")
                logger.warning(f"Publication export failed: {e}")

        # Export plots
        if self.config.export_plots:
            try:
                plot_files = self._export_plots(output_dir)
                result.generated_files.update(plot_files)
            except Exception as e:
                result.errors.append(f"Plot export failed: {e}")
                logger.warning(f"Plot export failed: {e}")

        # Export dashboard format
        if self.config.export_dashboard:
            try:
                dash_files = self._export_dashboard(output_dir, scenario_name)
                result.generated_files.update(dash_files)
            except Exception as e:
                result.errors.append(f"Dashboard export failed: {e}")
                logger.warning(f"Dashboard export failed: {e}")

        # Export multi-network results (if available)
        if self.config.export_multi_network and self.has_multi_network_results():
            try:
                multi_files = self._export_multi_network(output_dir)
                result.generated_files.update(multi_files)
            except Exception as e:
                result.errors.append(f"Multi-network export failed: {e}")
                logger.warning(f"Multi-network export failed: {e}")

        # Write export manifest
        manifest_path = os.path.join(output_dir, "export_manifest.json")
        self._write_manifest(manifest_path, result)
        result.generated_files["manifest"] = manifest_path

        logger.info(result.summary())

        return result

    def _export_data(self, output_dir: str) -> dict[str, str]:
        """Export raw data files."""
        data_dir = os.path.join(output_dir, "data")
        os.makedirs(data_dir, exist_ok=True)

        files = {}

        # CSV: Timeseries
        if self.config.export_csv:
            csv_path = os.path.join(data_dir, "timeseries.csv")
            self._timeseries_df.to_csv(csv_path, sep=";")
            files["data_timeseries_csv"] = csv_path
            logger.info(f"  Exported: {csv_path}")

        # JSON: Summary
        if self.config.export_json:
            json_path = os.path.join(data_dir, "summary.json")
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(self._summary_sections, f, indent=2, default=str)
            files["data_summary_json"] = json_path
            logger.info(f"  Exported: {json_path}")

            # Also export costs separately
            costs_path = os.path.join(data_dir, "costs.json")
            costs_data = {}
            if hasattr(self._active_result, "costs"):
                costs_data = dict(self._active_result.costs)
            with open(costs_path, "w", encoding="utf-8") as f:
                json.dump(costs_data, f, indent=2, default=str)
            files["data_costs_json"] = costs_path

        # Excel: Complete bundle
        if self.config.export_excel:
            try:
                xlsx_path = os.path.join(data_dir, "scenario.xlsx")
                self._export_excel_bundle(xlsx_path)
                files["data_scenario_xlsx"] = xlsx_path
                logger.info(f"  Exported: {xlsx_path}")
            except Exception as e:
                logger.warning(f"Excel export failed: {e}")

        return files

    def _export_publication(self, output_dir: str) -> dict[str, str]:
        """Export publication-ready materials."""
        pub_dir = os.path.join(output_dir, "publication")
        latex_dir = os.path.join(pub_dir, "latex")
        os.makedirs(latex_dir, exist_ok=True)

        files = {}

        try:
            from calion.io.publication_exporter import (
                export_kpi_summary,
                export_latex_tables,
            )

            # LaTeX tables
            latex_files = export_latex_tables(
                outdir=latex_dir,
                summary_sections=self._summary_sections,
                table_style=self.config.latex_style,
            )
            for name, path in latex_files.items():
                files[f"latex_{name}"] = path
                logger.info(f"  Exported: {path}")

            # KPI summary (JSON + CSV)
            kpi_path = export_kpi_summary(
                outdir=pub_dir,
                summary_sections=self._summary_sections,
            )
            files["publication_kpi_json"] = kpi_path
            logger.info(f"  Exported: {kpi_path}")

        except ImportError as e:
            logger.warning(f"Publication exporter not available: {e}")
            # Fallback: simple JSON export
            fallback_path = os.path.join(pub_dir, "kpi_summary.json")
            with open(fallback_path, "w", encoding="utf-8") as f:
                json.dump(self._summary_sections, f, indent=2, default=str)
            files["publication_kpi_json"] = fallback_path

        return files

    def _export_plots(self, output_dir: str) -> dict[str, str]:
        """Export publication-quality plots."""
        fig_dir = os.path.join(output_dir, "figures")
        os.makedirs(fig_dir, exist_ok=True)

        files = {}

        try:
            from calion.io.publication_plotter import (
                PublicationConfig,
                export_publication_plots,
            )

            # Apply publication style
            PublicationConfig.apply_style()

            # Get table from active result
            table = self._active_result.table
            series = dict(self._active_result.series) if hasattr(self._active_result, "series") else {}

            # Generate plots
            plot_files = export_publication_plots(
                outdir=fig_dir,
                table=table,
                series=series,
                summary_sections=self._summary_sections,
            )

            # plot_files returns dict[str, list[str]] (multiple formats per plot)
            for name, paths in plot_files.items():
                if isinstance(paths, list):
                    # Store first path as primary, but log all
                    if paths:
                        files[f"figure_{name}"] = paths[0]
                        for p in paths:
                            logger.info(f"  Exported: {p}")
                else:
                    files[f"figure_{name}"] = paths
                    logger.info(f"  Exported: {paths}")

        except ImportError as e:
            logger.warning(f"Publication plotter not available: {e}")
            # Fallback: basic matplotlib plots
            self._export_basic_plots(fig_dir, files)
        except Exception as e:
            logger.warning(f"Plot export failed: {e}")
            self._export_basic_plots(fig_dir, files)

        return files

    def _export_basic_plots(self, fig_dir: str, files: dict[str, str]) -> None:
        """Fallback: Export basic matplotlib plots."""
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt

            df = self._timeseries_df

            # Plot 1: Heat production overview
            _fig, ax = plt.subplots(figsize=(12, 6))

            # Find heat-related columns
            heat_cols = [c for c in df.columns if any(
                x in c.lower() for x in ["heat", "waerme", "th_mw", "q_"]
            )]

            if heat_cols:
                for col in heat_cols[:5]:  # Limit to 5 columns
                    ax.plot(df.index[:168], df[col].values[:168], label=col, alpha=0.8)

                ax.set_xlabel("Time")
                ax.set_ylabel("Thermal Power [MW]")
                ax.set_title("Heat Production Profile (First Week)")
                ax.legend(loc="upper right", fontsize=8)
                ax.grid(True, alpha=0.3)

            plot_path = os.path.join(fig_dir, "heat_profile.png")
            plt.savefig(plot_path, dpi=self.config.plot_dpi, bbox_inches="tight")
            plt.close()
            files["figure_heat_profile"] = plot_path
            logger.info(f"  Exported: {plot_path}")

            # Plot 2: Cost breakdown pie chart
            objective = self._summary_sections.get("objective", {})
            cost_items = []

            cost_keys = [
                ("Grid_energy_cost_EUR", "Electricity"),
                ("Fuel_cost_EUR", "Fuel"),
                ("Capex_cost_EUR", "Investment"),
                ("CO2_cost_EUR", "CO2"),
                ("Demand_charge_cost_EUR", "Demand Charge"),
            ]

            for key, label in cost_keys:
                value = objective.get(key, 0)
                if value and abs(float(value)) > 1:
                    cost_items.append((label, float(value)))

            if cost_items:
                _fig, ax = plt.subplots(figsize=(8, 8))
                labels, values = zip(*cost_items, strict=False)
                ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90)
                ax.set_title("Cost Breakdown")

                plot_path = os.path.join(fig_dir, "cost_breakdown.png")
                plt.savefig(plot_path, dpi=self.config.plot_dpi, bbox_inches="tight")
                plt.close()
                files["figure_cost_breakdown"] = plot_path
                logger.info(f"  Exported: {plot_path}")

        except ImportError:
            logger.warning("Matplotlib not available for fallback plots")

    def _export_dashboard(
        self,
        output_dir: str,
        scenario_name: str | None = None,
    ) -> dict[str, str]:
        """Export dashboard-compatible format."""
        dash_dir = os.path.join(output_dir, "dashboard")
        os.makedirs(dash_dir, exist_ok=True)

        files = {}

        try:
            from calion.io.notebook_helpers import save_workflow_run

            dash_path = save_workflow_run(
                self.workflow,
                name=scenario_name or "Exported Scenario",
                description=f"Unified export at {datetime.now().isoformat()}",
                save_dir=dash_dir,
            )
            files["dashboard_run_json"] = dash_path
            logger.info(f"  Exported: {dash_path}")

        except ImportError as e:
            logger.warning(f"Dashboard helpers not available: {e}")
            # Fallback: simple JSON
            fallback_path = os.path.join(dash_dir, "workflow_run.json")
            self._export_dashboard_fallback(fallback_path, scenario_name)
            files["dashboard_run_json"] = fallback_path

        return files

    def _export_dashboard_fallback(self, path: str, scenario_name: str | None) -> None:
        """Fallback dashboard export."""
        data = {
            "name": scenario_name or "Scenario",
            "timestamp": datetime.now().isoformat(),
            "mode": self._get_workflow_mode(),
            "summary": self._summary_sections,
            "timeseries_columns": list(self._timeseries_df.columns),
            "timeseries_length": len(self._timeseries_df),
        }

        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, default=str)

    def _export_multi_network(self, output_dir: str) -> dict[str, str]:
        """
        Export multi-network specific data and visualizations.

        Creates:
        - multi_network/
            ├── networks_summary.json
            ├── heat_exchangers.json
            ├── network_temperatures.csv
            └── figures/
                ├── multi_network_temperatures.png
                ├── heat_exchanger_operation.png
                └── network_comparison.png
        """
        from calion.io.exporters import (
            export_multi_network_data,
            export_multi_network_latex,
            export_multi_network_plots,
        )

        multi_dir = os.path.join(output_dir, "multi_network")
        fig_dir = os.path.join(multi_dir, "figures")
        os.makedirs(fig_dir, exist_ok=True)

        files: dict[str, str] = {}
        multi_data = self._multi_network_results

        logger.info("Exporting multi-network results...")

        files.update(export_multi_network_data(multi_dir, multi_data, self._timeseries_df))

        try:
            table = self._active_result.table
            series = dict(self._active_result.series) if hasattr(self._active_result, "series") else {}
            files.update(export_multi_network_plots(
                fig_dir, multi_data, self._timeseries_df,
                table, series, plot_dpi=self.config.plot_dpi,
            ))
        except Exception as e:
            logger.warning(f"Multi-network plots failed: {e}")

        try:
            files.update(export_multi_network_latex(multi_dir, multi_data))
        except Exception as e:
            logger.warning(f"Multi-network LaTeX export failed: {e}")

        return files

    def _export_excel_bundle(self, path: str) -> None:
        """Export complete Excel workbook."""
        try:
            from calion.io.exporter import export_scenario_bundle

            # Get directory from path
            outdir = os.path.dirname(path)

            # Build timeseries sections for export_scenario_bundle
            table = self._active_result.table
            series = self._active_result.series if hasattr(self._active_result, "series") else {}

            timeseries_sections = [
                {
                    "label": "Input",
                    "timestamps": list(table.index),
                    "series": OrderedDict((col, list(table[col])) for col in table.columns),
                },
                {
                    "label": "Results",
                    "timestamps": list(table.index),
                    "series": OrderedDict((name, list(values)) for name, values in series.items()),
                },
            ]

            # Use existing exporter with correct signature
            export_scenario_bundle(
                outdir,
                meta_sections=self._summary_sections,
                timeseries_sections=timeseries_sections,
                cost_sections={"costs": self._active_result.costs} if hasattr(self._active_result, "costs") else None,
            )
        except (ImportError, Exception) as e:
            logger.warning(f"export_scenario_bundle failed: {e}, using pandas fallback")
            # Fallback: simple pandas export
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                self._timeseries_df.to_excel(writer, sheet_name="Timeseries")

                # Summary as separate sheet
                summary_df = pd.DataFrame([
                    {"Section": sec, "Metric": key, "Value": val}
                    for sec, metrics in self._summary_sections.items()
                    for key, val in (metrics.items() if isinstance(metrics, dict) else [("value", metrics)])
                ])
                summary_df.to_excel(writer, sheet_name="Summary", index=False)

    def _generate_output_dir(self, scenario_name: str | None = None) -> str:
        """Generate output directory path."""
        stamp = time.strftime("%Y%m%d_%H%M%S")

        if scenario_name:
            slug = self._slugify(scenario_name)
        else:
            mode = self._get_workflow_mode()
            cfg = self.workflow.config if hasattr(self.workflow, "config") else {}
            title = cfg.get("scenario", {}).get("title", "scenario")
            slug = self._slugify(f"{mode}_{title}")

        return os.path.join("exports", f"{stamp}_{slug}")

    def _get_workflow_mode(self) -> str:
        """Determine workflow mode from results."""
        if self.workflow.mpc_result:
            return "MPC"
        elif self.workflow.rh_result:
            return "RH"
        elif self.workflow.pf_result:
            return "PF"
        return "UNKNOWN"

    def _write_manifest(self, path: str, result: ExportResult) -> None:
        """Write export manifest."""
        manifest = {
            "export_timestamp": datetime.now().isoformat(),
            "output_dir": result.output_dir,
            "workflow_mode": self._get_workflow_mode(),
            "config": {
                "export_excel": self.config.export_excel,
                "export_csv": self.config.export_csv,
                "export_json": self.config.export_json,
                "export_latex": self.config.export_latex,
                "export_plots": self.config.export_plots,
                "export_dashboard": self.config.export_dashboard,
            },
            "generated_files": result.generated_files,
            "errors": result.errors,
        }

        with open(path, "w", encoding="utf-8") as f:
            json.dump(manifest, f, indent=2)

    @staticmethod
    def _slugify(text: str) -> str:
        """Convert text to filesystem-safe slug."""
        import re
        text = str(text).lower()
        text = re.sub(r"[^\w\s-]", "", text)
        text = re.sub(r"[-\s]+", "_", text)
        return text.strip("_")[:50]


# =============================================================================
# Convenience functions
# =============================================================================

def export_workflow(
    workflow: WorkflowResult,
    output_dir: str | None = None,
    scenario_name: str | None = None,
    **config_kwargs,
) -> ExportResult:
    """
    Convenience function to export workflow results.

    Parameters
    ----------
    workflow : WorkflowResult
        The workflow result from run_workflow()
    output_dir : str, optional
        Output directory
    scenario_name : str, optional
        Scenario name for filenames
    **config_kwargs
        Additional configuration options passed to ExportConfig

    Returns
    -------
    ExportResult
        Export summary

    Example
    -------
    >>> from calion.run.rolling_horizon import run_workflow
    >>> from calion.io.unified_exporter import export_workflow
    >>>
    >>> workflow = run_workflow(["configs/my_scenario.yaml"])
    >>> result = export_workflow(workflow, scenario_name="My Analysis")
    >>> print(result.summary())
    """
    config = ExportConfig(**config_kwargs)
    exporter = UnifiedExporter(workflow, config)
    return exporter.export_all(output_dir, scenario_name)


def export_with_sensitivity(
    workflow: WorkflowResult,
    base_configs: list[str],
    output_dir: str | None = None,
    scenario_name: str | None = None,
    sensitivity_kwargs: dict[str, Any] | None = None,
) -> ExportResult:
    """
    Export workflow results with sensitivity analysis.

    Parameters
    ----------
    workflow : WorkflowResult
        The workflow result
    base_configs : List[str]
        Config files used for sensitivity analysis
    output_dir : str, optional
        Output directory
    scenario_name : str, optional
        Scenario name
    sensitivity_kwargs : dict, optional
        Additional kwargs for run_full_sensitivity_study

    Returns
    -------
    ExportResult
        Export summary including sensitivity files
    """
    # First, do the standard export
    config = ExportConfig(export_sensitivity=True)
    exporter = UnifiedExporter(workflow, config)
    result = exporter.export_all(output_dir, scenario_name)

    # Then run sensitivity analysis
    try:
        from calion.analysis.sensitivity_runner import run_full_sensitivity_study

        sens_dir = os.path.join(result.output_dir, "sensitivity")
        sens_kwargs = sensitivity_kwargs or {}

        logger.info("Running sensitivity analysis...")
        sens_result = run_full_sensitivity_study(
            base_configs=base_configs,
            output_dir=sens_dir,
            **sens_kwargs,
        )

        # Add sensitivity files to result
        for name, path in sens_result.generated_files.items():
            result.generated_files[f"sensitivity_{name}"] = path

        logger.info(sens_result.summary())

    except Exception as e:
        result.errors.append(f"Sensitivity analysis failed: {e}")
        logger.warning(f"Sensitivity analysis failed: {e}")

    return result
